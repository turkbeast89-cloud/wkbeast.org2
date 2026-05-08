from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Request, Form
from fastapi.responses import StreamingResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone
from io import BytesIO
import openpyxl
import asyncio
import time

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Proxy for outbound API calls (ViaBTC requires static IP whitelisting)
PROXY_URL = os.environ.get('PROXY_URL', '')

def _proxy():
    """Return proxy URL for aiohttp requests, or None if not set"""
    return PROXY_URL if PROXY_URL else None

# Simple cache for machine monitor (30 second TTL)
_machine_monitor_cache = {"data": None, "timestamp": 0}
CACHE_TTL = 30  # seconds

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class MachineType(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    monthly_fee: float
    daily_profit: float = 0.0  # Daily profit estimate from asicminervalue.com
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class MachineTypeCreate(BaseModel):
    name: str
    monthly_fee: float
    daily_profit: float = 0.0

class CustomerMachine(BaseModel):
    machine_type_id: str
    machine_name: str
    quantity: int

class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: str = ""
    machines: List[CustomerMachine] = []
    total_cost: float = 0  # Your cost
    total_fee: float = 0   # What customer pays
    status: str = "active"  # active, paused
    prepaid_months: int = 0  # Months paid in advance
    whatsapp_enabled: bool = True  # Whether this customer receives WhatsApp messages
    notes: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class CustomerCreate(BaseModel):
    name: str
    phone: str = ""
    machines: List[Dict] = []
    total_cost: float = 0
    status: str = "active"
    prepaid_months: int = 0
    notes: str = ""

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    machines: Optional[List[Dict]] = None
    total_cost: Optional[float] = None
    status: Optional[str] = None
    prepaid_months: Optional[int] = None
    notes: Optional[str] = None

class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    customer_name: str
    month: str  # Format: "2024-01"
    amount: float
    status: str = "unpaid"  # paid, unpaid, paused
    paid_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class PaymentUpdate(BaseModel):
    status: str
    amount: Optional[float] = None

class Repair(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    customer_name: str
    description: str  # What was repaired
    cost: float
    status: str = "unpaid"  # unpaid, paid
    paid_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Settings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "settings"
    message_template: str = ""
    whish_number: str = "03022005"
    usdt_address: str = "0x4e44e18349c4531f4463Fc49056b182C28C54877"
    team_name: str = "WKBeast Team"

class SettingsUpdate(BaseModel):
    message_template: Optional[str] = None
    whish_number: Optional[str] = None
    usdt_address: Optional[str] = None
    team_name: Optional[str] = None

# ==================== CUSTOMER PORTAL MODELS ====================

class CustomerAccount(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str  # Links to Customer
    username: str  # Customer name (lowercase, no spaces)
    password: str  # Last 4 digits of phone
    worker_name: str = ""  # ViaBTC worker name
    viabtc_api_key: str = ""  # Customer's own ViaBTC API key
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class CustomerAccountCreate(BaseModel):
    customer_id: str
    username: str
    password: str
    worker_name: str = ""
    viabtc_api_key: str = ""

class MaintenanceLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    machine_info: str = ""
    description: str
    date: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class MaintenanceLogCreate(BaseModel):
    customer_id: str
    machine_info: str = ""
    description: str

class FarmStats(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "farm_stats"
    machines_online: int = 2430
    machines_offline: int = 10
    total_hashrate: str = "850 TH/s"
    total_hashrate_by_coin: str = ""  # Format: "LTC:500 GH/s,KAS:350 TH/s"
    fluctuation: int = 5  # Random +/- range

class FarmStatsUpdate(BaseModel):
    machines_online: Optional[int] = None
    machines_offline: Optional[int] = None
    total_hashrate: Optional[str] = None
    total_hashrate_by_coin: Optional[str] = None
    fluctuation: Optional[int] = None

class ViaBTCSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "viabtc_settings"
    access_key: str = ""
    secret_key: str = ""
    enabled: bool = False

class MachineStatus(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    worker_name: str
    status: str = "online"  # online, offline
    hashrate: str = "0 TH/s"
    temperature: str = "0°C"
    uptime: str = "0h"
    last_updated: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# ==================== MACHINE TYPES ====================

@api_router.get("/machine-types", response_model=List[MachineType])
async def get_machine_types():
    types = await db.machine_types.find({}, {"_id": 0}).to_list(100)
    return types

@api_router.get("/server-ip")
async def get_server_ip():
    """Get the server's external IP for ViaBTC whitelisting"""
    import aiohttp
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get("https://api.ipify.org", timeout=5, proxy=_proxy()) as resp:
                ip = (await resp.text()).strip()
                return {"ip": ip, "message": f"Whitelist this IP in ViaBTC: {ip}", "proxy_enabled": bool(PROXY_URL)}
    except:
        return {"ip": "unknown", "message": "Could not fetch IP. Try: curl https://api.ipify.org"}

@api_router.post("/machine-types", response_model=MachineType)
async def create_machine_type(data: MachineTypeCreate):
    machine_type = MachineType(**data.model_dump())
    doc = machine_type.model_dump()
    await db.machine_types.insert_one(doc)
    return machine_type

@api_router.put("/machine-types/{type_id}", response_model=MachineType)
async def update_machine_type(type_id: str, data: MachineTypeCreate):
    result = await db.machine_types.find_one_and_update(
        {"id": type_id},
        {"$set": {"name": data.name, "monthly_fee": data.monthly_fee, "daily_profit": data.daily_profit}},
        return_document=True,
        projection={"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Machine type not found")
    return result

@api_router.delete("/machine-types/{type_id}")
async def delete_machine_type(type_id: str):
    result = await db.machine_types.delete_one({"id": type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Machine type not found")
    return {"message": "Deleted"}

# ==================== CUSTOMERS ====================

@api_router.get("/customers", response_model=List[Customer])
async def get_customers():
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    return customers

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer

@api_router.post("/customers", response_model=Customer)
async def create_customer(data: CustomerCreate):
    # Calculate total fee from machines
    machine_types = await db.machine_types.find({}, {"_id": 0}).to_list(100)
    machine_fee_map = {mt["id"]: mt["monthly_fee"] for mt in machine_types}
    machine_name_map = {mt["id"]: mt["name"] for mt in machine_types}
    
    total_fee = 0
    machines_with_names = []
    for m in data.machines:
        fee = machine_fee_map.get(m.get("machine_type_id"), 0)
        name = machine_name_map.get(m.get("machine_type_id"), "Unknown")
        qty = m.get("quantity", 1)
        total_fee += fee * qty
        machines_with_names.append({
            "machine_type_id": m.get("machine_type_id"),
            "machine_name": name,
            "quantity": qty
        })
    
    customer = Customer(
        name=data.name,
        phone=data.phone,
        machines=machines_with_names,
        total_cost=data.total_cost,
        total_fee=total_fee,
        status=data.status,
        prepaid_months=data.prepaid_months,
        notes=data.notes
    )
    doc = customer.model_dump()
    await db.customers.insert_one(doc)
    return customer

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, data: CustomerUpdate):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    # Recalculate total fee if machines changed
    if "machines" in update_data:
        machine_types = await db.machine_types.find({}, {"_id": 0}).to_list(100)
        machine_fee_map = {mt["id"]: mt["monthly_fee"] for mt in machine_types}
        machine_name_map = {mt["id"]: mt["name"] for mt in machine_types}
        
        total_fee = 0
        machines_with_names = []
        for m in update_data["machines"]:
            fee = machine_fee_map.get(m.get("machine_type_id"), 0)
            name = machine_name_map.get(m.get("machine_type_id"), "Unknown")
            qty = m.get("quantity", 1)
            total_fee += fee * qty
            machines_with_names.append({
                "machine_type_id": m.get("machine_type_id"),
                "machine_name": name,
                "quantity": qty
            })
        update_data["machines"] = machines_with_names
        update_data["total_fee"] = total_fee
    
    result = await db.customers.find_one_and_update(
        {"id": customer_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Customer not found")
    return result

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str):
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    # Also delete customer's payments
    await db.payments.delete_many({"customer_id": customer_id})
    return {"message": "Deleted"}

# ==================== PAYMENTS ====================

@api_router.get("/payments")
async def get_payments(month: Optional[str] = None):
    query = {}
    if month:
        query["month"] = month
    payments = await db.payments.find(query, {"_id": 0}).to_list(10000)
    return payments

@api_router.post("/payments/generate")
async def generate_monthly_payments(month: str):
    """Generate payment records for all active customers for a specific month"""
    customers = await db.customers.find({"status": "active"}, {"_id": 0}).to_list(1000)
    
    created = 0
    for customer in customers:
        # Check if payment already exists
        existing = await db.payments.find_one({
            "customer_id": customer["id"],
            "month": month
        })
        if not existing:
            # Check prepaid months
            prepaid = customer.get("prepaid_months", 0)
            status = "paid" if prepaid > 0 else "unpaid"
            
            payment = Payment(
                customer_id=customer["id"],
                customer_name=customer["name"],
                month=month,
                amount=customer["total_fee"],
                status=status
            )
            await db.payments.insert_one(payment.model_dump())
            
            # Decrease prepaid months if used
            if prepaid > 0:
                await db.customers.update_one(
                    {"id": customer["id"]},
                    {"$inc": {"prepaid_months": -1}}
                )
            created += 1
    
    return {"message": f"Generated {created} payment records for {month}"}

@api_router.put("/payments/{payment_id}")
async def update_payment(payment_id: str, data: PaymentUpdate):
    update_data = {"status": data.status}
    if data.status == "paid":
        update_data["paid_at"] = datetime.now(timezone.utc).isoformat()
    if data.amount is not None:
        update_data["amount"] = data.amount
    
    result = await db.payments.find_one_and_update(
        {"id": payment_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Auto-send WhatsApp confirmation when marked as paid
    if data.status == "paid":
        try:
            bot_settings = await db.bot_settings.find_one({"id": "bot_settings"}, {"_id": 0})
            if bot_settings and bot_settings.get("auto_reminders_enabled"):
                customer = await db.customers.find_one({"id": result.get("customer_id")}, {"_id": 0})
                if customer and customer.get("phone") and customer.get("status") != "paused":
                    phone_clean = format_phone_whatsapp(customer["phone"])
                    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
                    team = settings.get("team_name", "WKBeast Team") if settings else "WKBeast Team"
                    month = result.get("month", "")
                    
                    message = f"""✅ Payment Confirmed

Dear {customer.get('name', 'Customer')},

Thank you! We have reviewed and confirmed your payment for {month}.

Your machines will continue to run without interruption. We appreciate your timely payment.

Best regards,
{team} 🐺💼

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""
                    
                    client_tw = get_twilio_client()
                    if client_tw:
                        client_tw.messages.create(body=message, from_=get_whatsapp_from(), to=f"whatsapp:{phone_clean}")
                        await db.whatsapp_logs.insert_one({
                            "customer_id": result.get("customer_id", ""),
                            "customer_name": customer.get("name", ""),
                            "phone": phone_clean,
                            "message": message,
                            "type": "payment_confirmed",
                            "direction": "outbound",
                            "status": "sent",
                            "created_at": datetime.now(timezone.utc).isoformat()
                        })
        except:
            pass  # Don't fail the payment update if WhatsApp fails
    
    return result

@api_router.post("/payments/bulk-status")
async def bulk_update_payment_status(payment_ids: List[str], status: str):
    update_data = {"status": status}
    if status == "paid":
        update_data["paid_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.payments.update_many(
        {"id": {"$in": payment_ids}},
        {"$set": update_data}
    )
    return {"modified": result.modified_count}


@api_router.get("/payments/overdue")
async def get_overdue_payments():
    """Get all unpaid payments from previous months grouped by customer"""
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    # Find all unpaid payments (current month and older)
    unpaid = await db.payments.find({"status": "unpaid"}, {"_id": 0}).to_list(10000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    customer_map = {c["id"]: c for c in customers}
    
    overdue = []
    for p in unpaid:
        if p.get("month", "") < current_month:
            customer = customer_map.get(p.get("customer_id"), {})
            overdue.append({
                "payment_id": p.get("id", ""),
                "customer_id": p.get("customer_id", ""),
                "customer_name": p.get("customer_name", customer.get("name", "Unknown")),
                "phone": customer.get("phone", ""),
                "amount": p.get("amount", 0),
                "month": p.get("month", ""),
                "status": customer.get("status", "active")
            })
    
    # Group by customer
    by_customer = {}
    for o in overdue:
        cname = o["customer_name"]
        if cname not in by_customer:
            by_customer[cname] = {
                "customer_name": cname,
                "customer_id": o["customer_id"],
                "phone": o["phone"],
                "status": o["status"],
                "total_owed": 0,
                "months": []
            }
        by_customer[cname]["total_owed"] += o["amount"]
        by_customer[cname]["months"].append({"month": o["month"], "amount": o["amount"], "payment_id": o["payment_id"]})
    
    # Sort by total owed descending
    result = sorted(by_customer.values(), key=lambda x: x["total_owed"], reverse=True)
    
    return {"overdue_customers": result, "total_overdue": sum(c["total_owed"] for c in result), "total_customers": len(result)}




# ==================== MACHINE DATA SYNC (from PC app) ====================

@api_router.post("/machine-data/push")
async def push_machine_data(data: dict):
    """Receive machine data from the PC monitoring app"""
    api_key = data.get("api_key", "")
    
    # Simple API key auth
    expected_key = os.environ.get("MACHINE_SYNC_KEY", "wkbeast2025sync")
    if api_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid API key")
    
    machines = data.get("machines", [])
    farm = data.get("farm", "Main Farm")
    if not machines:
        return {"success": False, "error": "No machine data"}
    
    # Get list of deleted/hidden IPs (don't re-add them)
    hidden = await db.machine_hidden.find({}, {"_id": 0}).to_list(10000)
    hidden_ips = set(h.get("ip", "") for h in hidden)
    
    # Get all manually renamed machines in one query
    renamed = await db.machine_live_data.find({"manually_renamed": True}, {"_id": 0, "ip": 1}).to_list(10000)
    renamed_ips = set(m.get("ip", "") for m in renamed)
    
    synced = 0
    ops = []
    for m in machines:
        ip = m.get("ip", "")
        if not ip or ip in hidden_ips:
            continue

        # Build update dict but DO NOT overwrite worker_name/model with empty values.
        # When a miner goes offline, MineFleet often pushes with blank worker_name,
        # which would erase the customer-tag and make the row vanish from "My Customers".
        new_worker_name = (m.get("worker_name") or "").strip()
        new_model = (m.get("model") or "").strip()

        update_data = {
            "ip": ip,
            "hashrate": m.get("hashrate", 0),
            "temperature": m.get("temperature", 0),
            "fan_speed": m.get("fan_speed", 0),
            "power": m.get("power", 0),
            "status": m.get("status") or "online",
            "uptime": m.get("uptime", ""),
            "pool": m.get("pool", ""),
            "farm": m.get("farm", farm),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }

        # Track when the machine was last seen ONLINE so we know its real "last alive" time
        if update_data["status"] == "online":
            update_data["last_online_at"] = update_data["updated_at"]

        if ip not in renamed_ips and new_worker_name:
            update_data["worker_name"] = new_worker_name
        if new_model:
            update_data["model"] = new_model

        from pymongo import UpdateOne
        ops.append(UpdateOne({"ip": ip}, {"$set": update_data}, upsert=True))
        synced += 1

    if ops:
        await db.machine_live_data.bulk_write(ops)

    return {"success": True, "synced": synced}

@api_router.get("/machine-data/live")
async def get_live_machine_data(filter_customers: bool = True):
    """Get live machine data. filter_customers=true shows only customer machines, false shows all."""
    machines = await db.machine_live_data.find({}, {"_id": 0}).to_list(10000)

    # Get all customer worker names
    accounts = await db.customer_accounts.find({}, {"_id": 0, "worker_name": 1}).to_list(10000)
    customer_workers = set()
    for acc in accounts:
        wn = acc.get("worker_name", "").lower().strip()
        if wn:
            customer_workers.add(wn)

    # Get currently-saved "original worker" memory keyed by IP (from a wallet switch)
    switch_data = await db.wallet_switch.find_one({"id": "wallet_switch"}, {"_id": 0})
    originals_by_ip = {}
    switched_active = bool(switch_data and switch_data.get("switched"))
    if switch_data:
        for o in (switch_data.get("originals") or []):
            originals_by_ip[o.get("ip", "")] = o.get("original_worker", "")

    # Mark each machine as customer or not + attach original_worker memory.
    # Also auto-flip status to "offline" if the machine hasn't been updated in >10 minutes
    # (PC sync runs every 2 min; >10 min means it's been missing for 5+ cycles).
    now = datetime.now(timezone.utc)
    STALE_THRESHOLD_SECONDS = 600  # 10 minutes

    for m in machines:
        worker = m.get("worker_name", "").lower().strip()
        is_customer = False
        if worker:
            for cw in customer_workers:
                if cw in worker or worker in cw:
                    is_customer = True
                    break
        m["is_customer"] = is_customer

        # Compute staleness so the dashboard can show "offline" even if the row was
        # never refreshed with status=offline (e.g. MineFleet dropped it from config).
        updated_at_str = m.get("updated_at", "")
        seconds_since_update = None
        if updated_at_str:
            try:
                updated_dt = datetime.fromisoformat(updated_at_str.replace("Z", "+00:00"))
                seconds_since_update = (now - updated_dt).total_seconds()
            except Exception:
                seconds_since_update = None

        if seconds_since_update is not None and seconds_since_update > STALE_THRESHOLD_SECONDS:
            m["status"] = "offline"
            m["stale"] = True
            m["seconds_since_update"] = int(seconds_since_update)
        else:
            m["stale"] = False

        # Attach original_worker if a switch is active AND the original differs from the current worker
        orig = originals_by_ip.get(m.get("ip", ""), "").strip()
        if switched_active and orig and orig.lower() != (m.get("worker_name", "") or "").lower():
            m["original_worker"] = orig

    if filter_customers:
        return [m for m in machines if m["is_customer"]]

    return machines


@api_router.post("/machine-data/clear-original")
async def clear_original_for_ip(data: dict):
    """Forget the saved 'original worker' memory for one specific IP.
    Use this when you want to keep the temporary worker permanent (no restore).
    Body: {ip}
    If clearing this leaves the originals list empty, the wallet switch is auto-deactivated
    so the 'Restore Original Workers' button disappears.
    """
    ip = (data.get("ip") or "").strip()
    if not ip:
        raise HTTPException(status_code=400, detail="ip is required")

    switch_data = await db.wallet_switch.find_one({"id": "wallet_switch"}, {"_id": 0})
    if not switch_data:
        return {"success": True, "removed": 0, "switch_active": False}

    originals = switch_data.get("originals") or []
    new_originals = [o for o in originals if o.get("ip") != ip]
    removed = len(originals) - len(new_originals)

    update_doc = {"originals": new_originals}
    # Auto-deactivate the switch when no originals remain
    if not new_originals:
        update_doc["switched"] = False
        update_doc["restored_at"] = datetime.now(timezone.utc).isoformat()

    await db.wallet_switch.update_one(
        {"id": "wallet_switch"},
        {"$set": update_doc}
    )

    return {"success": True, "removed": removed, "switch_active": bool(new_originals)}

@api_router.get("/machine-data/commands")
async def get_pending_commands(api_key: str = "", farm: str = ""):
    """PC app polls this to check for commands to execute"""
    expected_key = os.environ.get("MACHINE_SYNC_KEY", "wkbeast2025sync")
    if api_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid API key")
    
    query = {"status": "pending"}
    if farm:
        query["farm"] = farm
    commands = await db.machine_commands.find(query, {"_id": 0}).to_list(100)
    return commands

@api_router.post("/machine-data/command")
async def create_machine_command(data: dict):
    """Create a command for the PC app to execute (reboot, change worker, etc.)"""
    ip = data.get("ip", "")
    # Look up which farm this machine belongs to
    machine = await db.machine_live_data.find_one({"ip": ip}, {"_id": 0})
    farm = machine.get("farm", "Main Farm") if machine else "Main Farm"
    
    cmd = {
        "id": str(uuid.uuid4()),
        "ip": ip,
        "action": data.get("action", "reboot"),  # reboot, change_worker, change_pool
        "params": data.get("params", {}),
        "farm": farm,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": None,
        "result": None
    }
    await db.machine_commands.insert_one(cmd)
    return {"success": True, "command_id": cmd["id"]}

@api_router.post("/machine-data/command-done")
async def complete_command(data: dict):
    """PC app reports a command was executed"""
    api_key = data.get("api_key", "")
    expected_key = os.environ.get("MACHINE_SYNC_KEY", "wkbeast2025sync")
    if api_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid API key")
    
    await db.machine_commands.update_one(
        {"id": data.get("command_id")},
        {"$set": {
            "status": "completed",
            "result": data.get("result", "OK"),
            "completed_at": datetime.now(timezone.utc).isoformat()
        }}
    )

@api_router.post("/machine-data/rename")
async def rename_machine(data: dict):
    """Rename a machine's worker name"""
    ip = data.get("ip", "")
    new_name = data.get("worker_name", "")
    if not ip or not new_name:
        raise HTTPException(status_code=400, detail="IP and worker_name required")
    
    result = await db.machine_live_data.update_one({"ip": ip}, {"$set": {"worker_name": new_name, "manually_renamed": True}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Machine not found")
    return {"success": True}

@api_router.delete("/machine-data/{ip}")
async def delete_machine(ip: str):
    """Remove a machine from live data and prevent it from coming back"""
    await db.machine_live_data.delete_one({"ip": ip})
    # Add to hidden list so sync doesn't re-add it
    await db.machine_hidden.update_one({"ip": ip}, {"$set": {"ip": ip, "hidden_at": datetime.now(timezone.utc).isoformat()}}, upsert=True)
    return {"success": True}



@api_router.post("/machine-data/switch-workers")
async def switch_workers(data: dict):
    """Switch workers on selected or all machines to a new wallet"""
    new_worker = data.get("new_worker", "")
    pool_url = data.get("pool_url", "")
    machine_ips = data.get("ips", [])  # Empty = all customer machines
    
    if not new_worker:
        raise HTTPException(status_code=400, detail="new_worker is required")
    
    # If no specific IPs, get all customer machines
    if not machine_ips:
        accounts = await db.customer_accounts.find({}, {"_id": 0, "worker_name": 1}).to_list(10000)
        customer_workers = set(a.get("worker_name", "").lower().strip() for a in accounts if a.get("worker_name"))
        
        all_machines = await db.machine_live_data.find({}, {"_id": 0}).to_list(10000)
        for m in all_machines:
            worker = m.get("worker_name", "").lower().strip()
            for cw in customer_workers:
                if cw in worker or worker in cw:
                    machine_ips.append(m["ip"])
                    break
    
    # Save original workers before switching (for restore)
    # IMPORTANT: if a switch is already active, do NOT overwrite the saved originals
    # otherwise restore would target the already-switched values instead of the real originals.
    existing_switch = await db.wallet_switch.find_one({"id": "wallet_switch"}, {"_id": 0})
    already_switched = bool(existing_switch and existing_switch.get("switched"))
    existing_originals_by_ip = {}
    if already_switched:
        for o in (existing_switch.get("originals") or []):
            existing_originals_by_ip[o.get("ip", "")] = o.get("original_worker", "")

    originals = []
    commands_created = 0

    for ip in machine_ips:
        machine = await db.machine_live_data.find_one({"ip": ip}, {"_id": 0})
        if machine:
            # Preserve previously-saved original if a switch is already active for this IP
            if ip in existing_originals_by_ip:
                originals.append({"ip": ip, "original_worker": existing_originals_by_ip[ip]})
            else:
                originals.append({"ip": ip, "original_worker": machine.get("worker_name", "")})

            # Look up farm for this machine
            farm = machine.get("farm", "Main Farm")
            
            cmd = {
                "id": str(uuid.uuid4()),
                "ip": ip,
                "action": "change_worker",
                "params": {"worker_name": new_worker, "pool_url": pool_url},
                "farm": farm,
                "status": "pending",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "completed_at": None,
                "result": None
            }
            await db.machine_commands.insert_one(cmd)
            commands_created += 1
    
    # Save originals for restore
    await db.wallet_switch.update_one(
        {"id": "wallet_switch"},
        {"$set": {
            "id": "wallet_switch",
            "switched": True,
            "new_worker": new_worker,
            "originals": originals,
            "switched_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"success": True, "commands": commands_created, "machines": len(machine_ips)}

@api_router.post("/machine-data/restore-workers")
async def restore_workers():
    """Restore all workers back to their original wallet"""
    switch_data = await db.wallet_switch.find_one({"id": "wallet_switch"}, {"_id": 0})
    if not switch_data or not switch_data.get("switched"):
        return {"success": False, "error": "No active wallet switch to restore"}
    
    originals = switch_data.get("originals", [])
    commands_created = 0
    
    for item in originals:
        ip = item["ip"]
        original_worker = item["original_worker"]
        
        machine = await db.machine_live_data.find_one({"ip": ip}, {"_id": 0})
        farm = machine.get("farm", "Main Farm") if machine else "Main Farm"
        
        cmd = {
            "id": str(uuid.uuid4()),
            "ip": ip,
            "action": "change_worker",
            "params": {"worker_name": original_worker},
            "farm": farm,
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "completed_at": None,
            "result": None
        }
        await db.machine_commands.insert_one(cmd)
        commands_created += 1
    
    # Mark as restored
    await db.wallet_switch.update_one(
        {"id": "wallet_switch"},
        {"$set": {"switched": False, "restored_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"success": True, "commands": commands_created, "restored": len(originals)}

@api_router.get("/machine-data/switch-status")
async def get_switch_status():
    """Check if workers are currently switched"""
    switch_data = await db.wallet_switch.find_one({"id": "wallet_switch"}, {"_id": 0})
    if not switch_data:
        return {"switched": False}
    return switch_data


    return {"success": True}


# ==================== REPAIRS ====================

@api_router.get("/repairs")
async def get_repairs(customer_id: str = None):
    """Get all repairs, optionally filtered by customer"""
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    repairs = await db.repairs.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    return repairs

@api_router.post("/repairs")
async def create_repair(data: dict):
    """Create a repair record for a customer"""
    customer = await db.customers.find_one({"id": data.get("customer_id")}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    repair = Repair(
        customer_id=data["customer_id"],
        customer_name=customer.get("name", ""),
        description=data.get("description", ""),
        cost=float(data.get("cost", 0))
    )
    await db.repairs.insert_one(repair.model_dump())
    return repair.model_dump()

@api_router.put("/repairs/{repair_id}")
async def update_repair(repair_id: str, data: dict):
    """Update a repair - mark as paid or edit"""
    update_data = {}
    if "status" in data:
        update_data["status"] = data["status"]
        if data["status"] == "paid":
            update_data["paid_at"] = datetime.now(timezone.utc).isoformat()
    if "description" in data:
        update_data["description"] = data["description"]
    if "cost" in data:
        update_data["cost"] = float(data["cost"])
    
    result = await db.repairs.find_one_and_update(
        {"id": repair_id},
        {"$set": update_data},
        return_document=True,
        projection={"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Repair not found")
    
    # Send WhatsApp confirmation if marked as paid
    if data.get("status") == "paid":
        try:
            customer = await db.customers.find_one({"id": result["customer_id"]}, {"_id": 0})
            if customer and customer.get("phone") and customer.get("whatsapp_enabled", True):
                client_tw = get_twilio_client()
                if client_tw:
                    phone_clean = format_phone_whatsapp(customer["phone"])
                    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
                    team = settings.get("team_name", "WKBeast Team") if settings else "WKBeast Team"
                    msg = f"""✅ Repair Payment Confirmed

Dear {customer.get('name', 'Customer')},

Your repair fee of ${result['cost']} for "{result['description']}" has been marked as paid.

Thank you!
{team} 🐺💼

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""
                    client_tw.messages.create(body=msg, from_=get_whatsapp_from(), to=f"whatsapp:{phone_clean}")
        except:
            pass
    
    return result

@api_router.delete("/repairs/{repair_id}")
async def delete_repair(repair_id: str):
    """Delete a repair record"""
    result = await db.repairs.delete_one({"id": repair_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Repair not found")
    return {"success": True}


# ==================== STATISTICS ====================

@api_router.get("/stats")
async def get_stats():
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    payments = await db.payments.find({}, {"_id": 0}).to_list(10000)
    
    total_customers = len(customers)
    active_customers = len([c for c in customers if c.get("status") == "active"])
    paused_customers = len([c for c in customers if c.get("status") == "paused"])
    
    total_monthly_revenue = sum(c.get("total_fee", 0) for c in customers if c.get("status") == "active")
    total_monthly_cost = sum(c.get("total_cost", 0) for c in customers if c.get("status") == "active")
    monthly_profit = total_monthly_revenue - total_monthly_cost
    
    # Payment stats by month
    monthly_stats = {}
    for p in payments:
        month = p.get("month", "Unknown")
        if month not in monthly_stats:
            monthly_stats[month] = {"paid": 0, "unpaid": 0, "paused": 0, "total": 0}
        
        status = p.get("status", "unpaid")
        amount = p.get("amount", 0)
        monthly_stats[month][status] = monthly_stats[month].get(status, 0) + amount
        monthly_stats[month]["total"] += amount
    
    # Total collected vs pending
    total_collected = sum(p.get("amount", 0) for p in payments if p.get("status") == "paid")
    total_pending = sum(p.get("amount", 0) for p in payments if p.get("status") == "unpaid")
    
    # Machine breakdown (normalize variants like L9-275 → L9)
    machine_counts = {}
    for c in customers:
        for m in c.get("machines", []):
            name = m.get("machine_name", "Unknown")
            # Normalize: L9-275, L9-260 → L9
            import re
            base_name = re.sub(r'[-_]\d+$', '', name.upper())
            qty = m.get("quantity", 1)
            machine_counts[base_name] = machine_counts.get(base_name, 0) + qty
    
    return {
        "total_customers": total_customers,
        "active_customers": active_customers,
        "paused_customers": paused_customers,
        "total_monthly_revenue": total_monthly_revenue,
        "total_monthly_cost": total_monthly_cost,
        "monthly_profit": monthly_profit,
        "profit_margin": (monthly_profit / total_monthly_revenue * 100) if total_monthly_revenue > 0 else 0,
        "total_collected": total_collected,
        "total_pending": total_pending,
        "monthly_stats": monthly_stats,
        "machine_counts": machine_counts
    }

# ==================== SETTINGS ====================

@api_router.get("/settings", response_model=Settings)
async def get_settings():
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    if not settings:
        # Create default settings
        default = Settings(
            message_template="""📢 Dear Valued Customer,

This is a kind reminder to please settle your hosting fees before the 2nd of each month, as we have major financial obligations to cover by that date.

Unlike other farms, we don't request payments 6 months in advance — but we kindly ask for your cooperation in making the payment on time each month.

🔹 {month} Hosting Fee: ${amount}
🔹 Payment Options:
• Whish: {whish}
• USDT (BEP20 Network): {usdt}

Thank you for your continued trust and support. Your timely payment helps us keep everything running smoothly.

Warm regards,
{team} 🐺💼"""
        )
        await db.settings.insert_one(default.model_dump())
        return default
    return settings

@api_router.put("/settings", response_model=Settings)
async def update_settings(data: SettingsUpdate):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    result = await db.settings.find_one_and_update(
        {"id": "settings"},
        {"$set": update_data},
        upsert=True,
        return_document=True,
        projection={"_id": 0}
    )
    return result

# ==================== WHATSAPP ====================

@api_router.post("/whatsapp/generate-links")
async def generate_whatsapp_links(month: str, include_paid: bool = False):
    """Generate WhatsApp links for unpaid customers"""
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    if not settings:
        settings = {
            "message_template": "Payment reminder for {month}: ${amount}",
            "whish_number": "03022005",
            "usdt_address": "0x4e44e18349c4531f4463Fc49056b182C28C54877",
            "team_name": "WKBeast Team"
        }
    
    # Get payments for the month
    query = {"month": month}
    if not include_paid:
        query["status"] = "unpaid"
    
    payments = await db.payments.find(query, {"_id": 0}).to_list(1000)
    
    # Get customer phone numbers
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    customer_map = {c["id"]: c for c in customers}
    
    links = []
    for payment in payments:
        customer = customer_map.get(payment["customer_id"])
        if not customer or not customer.get("phone"):
            continue
        
        # Skip paused customers
        if customer.get("status") == "paused":
            continue
        
        # Format message
        message = settings["message_template"].format(
            month=month,
            amount=payment["amount"],
            whish=settings["whish_number"],
            usdt=settings["usdt_address"],
            team=settings["team_name"]
        )
        
        # Clean phone number for wa.me link (needs number without +)
        phone_formatted = format_phone_whatsapp(customer["phone"])
        phone = phone_formatted.lstrip("+")
        
        # Create WhatsApp link
        import urllib.parse
        encoded_message = urllib.parse.quote(message)
        link = f"https://wa.me/{phone}?text={encoded_message}"
        
        links.append({
            "customer_id": customer["id"],
            "customer_name": customer["name"],
            "phone": customer["phone"],
            "amount": payment["amount"],
            "link": link,
            "payment_id": payment["id"]
        })
    
    return links

# ==================== EXCEL IMPORT/EXPORT ====================

def normalize_machine_name(name):
    """Normalize machine names like L9-275, l9-260, L9 to base type L9"""
    import re
    name = name.strip().upper()
    # Remove variants like -275, -260, -250
    name = re.sub(r'[-_]\d+$', '', name)
    # Common normalizations
    name_map = {
        'KS5L': 'KS5PRO',
        'KS5': 'KS5PRO',
    }
    return name_map.get(name, name)

@api_router.post("/import/excel")
async def import_excel(file: UploadFile = File(...)):
    """Import customers from Excel file"""
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active
    
    # Get machine types for mapping
    machine_types = await db.machine_types.find({}, {"_id": 0}).to_list(100)
    machine_name_map = {mt["name"].upper(): mt for mt in machine_types}
    
    imported = 0
    errors = []
    
    # Get header row to find column indices
    headers = [str(cell).strip().lower() if cell else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    
    # Find column indices
    name_idx = next((i for i, h in enumerate(headers) if 'name' in h), 0)
    phone_idx = next((i for i, h in enumerate(headers) if 'phone' in h), 1)
    machines_idx = next((i for i, h in enumerate(headers) if 'machine' in h), 2)
    cost_idx = next((i for i, h in enumerate(headers) if 'cost' in h or 'your' in h), 3)
    fee_idx = next((i for i, h in enumerate(headers) if 'fee' in h or 'customer' in h), 4)
    status_idx = next((i for i, h in enumerate(headers) if 'status' in h), 5)
    prepaid_idx = next((i for i, h in enumerate(headers) if 'prepaid' in h), 6)
    
    # Skip header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row or not row[name_idx]:  # Skip empty rows
                continue
            
            name = str(row[name_idx]).strip()
            phone = str(row[phone_idx]).strip() if len(row) > phone_idx and row[phone_idx] else ""
            machines_str = str(row[machines_idx]).strip() if len(row) > machines_idx and row[machines_idx] else ""
            cost = float(row[cost_idx]) if len(row) > cost_idx and row[cost_idx] else 0
            fee = float(row[fee_idx]) if len(row) > fee_idx and row[fee_idx] else 0
            status = str(row[status_idx]).strip().lower() if len(row) > status_idx and row[status_idx] else "active"
            prepaid = int(row[prepaid_idx]) if len(row) > prepaid_idx and row[prepaid_idx] else 0
            
            # Parse machines string like "1x L9", "2x L9, 1x L7", "9x L9, 1x L1", "3x L9-275"
            machines = []
            total_fee = 0
            
            if machines_str:
                import re
                # Find patterns like "1x L9", "2x L9-275", "1x ks5pro"
                parts = re.findall(r'(\d+)\s*x\s*([a-zA-Z0-9\-_]+)', machines_str, re.IGNORECASE)
                
                machine_totals = {}  # Aggregate same machine types
                
                for qty_str, machine_name in parts:
                    qty = int(qty_str) if qty_str else 1
                    normalized_name = normalize_machine_name(machine_name)
                    
                    if normalized_name in machine_totals:
                        machine_totals[normalized_name] += qty
                    else:
                        machine_totals[normalized_name] = qty
                
                for normalized_name, qty in machine_totals.items():
                    if normalized_name in machine_name_map:
                        mt = machine_name_map[normalized_name]
                        machines.append({
                            "machine_type_id": mt["id"],
                            "machine_name": mt["name"],
                            "quantity": qty
                        })
                        total_fee += mt["monthly_fee"] * qty
            
            # Use provided fee if machines couldn't be parsed
            if total_fee == 0:
                total_fee = fee
            
            # Normalize status
            if status in ['paused', 'pause', 'inactive', 'off']:
                status = 'paused'
            else:
                status = 'active'
            
            customer = Customer(
                name=name,
                phone=phone,
                machines=machines,
                total_cost=cost,
                total_fee=total_fee,
                status=status,
                prepaid_months=prepaid
            )
            
            await db.customers.insert_one(customer.model_dump())
            imported += 1
            
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return {
        "imported": imported,
        "errors": errors
    }

@api_router.get("/export/excel")
async def export_excel():
    """Export customers to Excel file"""
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Headers
    ws.append(["Name", "Phone", "Machines", "Your Cost", "Customer Fee", "Status", "Prepaid Months", "Notes"])
    
    for customer in customers:
        machines_str = ", ".join([
            f"{m.get('quantity', 1)}x {m.get('machine_name', 'Unknown')}" 
            for m in customer.get("machines", [])
        ])
        ws.append([
            customer.get("name", ""),
            customer.get("phone", ""),
            machines_str,
            customer.get("total_cost", 0),
            customer.get("total_fee", 0),
            customer.get("status", "active"),
            customer.get("prepaid_months", 0),
            customer.get("notes", "")
        ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customers.xlsx"}
    )

@api_router.get("/export/full-backup")
async def export_full_backup():
    """Export ALL database data as a single Excel file with multiple sheets"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # 1. Customers sheet
    ws = wb.create_sheet("Customers")
    ws.append(["Name", "Phone", "Machines", "Your Cost", "Customer Fee", "Status", "Prepaid Months", "Notes", "Created At"])
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    for c in customers:
        machines_str = ", ".join([f"{m.get('quantity',1)}x {m.get('machine_name','?')}" for m in c.get("machines", [])])
        ws.append([c.get("name",""), c.get("phone",""), machines_str, c.get("total_cost",0), c.get("total_fee",0), c.get("status",""), c.get("prepaid_months",0), c.get("notes",""), c.get("created_at","")])
    
    # 2. Payments sheet
    ws2 = wb.create_sheet("Payments")
    ws2.append(["Customer Name", "Month", "Amount", "Status", "Paid At", "Created At"])
    payments = await db.payments.find({}, {"_id": 0}).to_list(50000)
    for p in payments:
        ws2.append([p.get("customer_name",""), p.get("month",""), p.get("amount",0), p.get("status",""), p.get("paid_at",""), p.get("created_at","")])
    
    # 3. Customer Accounts (Portal) sheet
    ws3 = wb.create_sheet("Customer Accounts")
    ws3.append(["Username", "Password", "Worker Name", "ViaBTC API Key", "Watcher Key", "Created At"])
    accounts = await db.customer_accounts.find({}, {"_id": 0}).to_list(10000)
    for a in accounts:
        ws3.append([a.get("username",""), a.get("password",""), a.get("worker_name",""), a.get("viabtc_api_key",""), a.get("watcher_key",""), a.get("created_at","")])
    
    # 4. Machine Types sheet
    ws4 = wb.create_sheet("Machine Types")
    ws4.append(["Name", "Monthly Fee", "Daily Profit"])
    mtypes = await db.machine_types.find({}, {"_id": 0}).to_list(100)
    for mt in mtypes:
        ws4.append([mt.get("name",""), mt.get("monthly_fee",0), mt.get("daily_profit",0)])
    
    # 5. Maintenance Logs sheet
    ws5 = wb.create_sheet("Maintenance Logs")
    ws5.append(["Customer ID", "Machine Info", "Description", "Date"])
    logs = await db.maintenance_logs.find({}, {"_id": 0}).to_list(10000)
    for l in logs:
        ws5.append([l.get("customer_id",""), l.get("machine_info",""), l.get("description",""), l.get("date","")])
    
    # 5b. Repairs sheet
    ws5b = wb.create_sheet("Repairs")
    ws5b.append(["Customer Name", "Description", "Cost", "Status", "Paid At", "Created At"])
    repairs = await db.repairs.find({}, {"_id": 0}).to_list(10000)
    for r in repairs:
        ws5b.append([r.get("customer_name",""), r.get("description",""), r.get("cost",0), r.get("status",""), r.get("paid_at",""), r.get("created_at","")])
    
    # 6. Settings sheet
    ws6 = wb.create_sheet("Settings")
    ws6.append(["Key", "Value"])
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    if settings:
        for k, v in settings.items():
            if k != "id":
                ws6.append([k, str(v)])
    
    # 7. ViaBTC Settings sheet
    ws7 = wb.create_sheet("ViaBTC Settings")
    ws7.append(["Key", "Value"])
    vbtc = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    if vbtc:
        for k, v in vbtc.items():
            if k != "id" and k != "secret_key":
                ws7.append([k, str(v)])
        if vbtc.get("secret_key"):
            ws7.append(["secret_key", "****" + vbtc["secret_key"][-4:]])
    
    # 8. Farm Stats sheet
    ws8 = wb.create_sheet("Farm Stats")
    ws8.append(["Key", "Value"])
    fstats = await db.farm_stats.find_one({"id": "farm_stats"}, {"_id": 0})
    if fstats:
        for k, v in fstats.items():
            if k != "id":
                ws8.append([k, str(v)])
    
    # Style all sheets - bold headers
    from openpyxl.styles import Font
    bold = Font(bold=True)
    for sheet in wb.sheetnames:
        for cell in wb[sheet][1]:
            cell.font = bold
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from datetime import datetime, timezone
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=wkbeast_full_backup_{timestamp}.xlsx"}
    )

@api_router.post("/import/full-backup")
async def import_full_backup(file: UploadFile = File(...), mode: str = "merge"):
    """Import a full backup Excel file. mode=merge (add/update) or mode=replace (wipe and restore)"""
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    
    results = {}
    
    try:
        # If replace mode, clear all collections first
        if mode == "replace":
            await db.customers.delete_many({})
            await db.payments.delete_many({})
            await db.customer_accounts.delete_many({})
            await db.machine_types.delete_many({})
            await db.maintenance_logs.delete_many({})
            results["mode"] = "replace (all data cleared first)"
        else:
            results["mode"] = "merge (added to existing data)"
        
        # 1. Machine Types
        if "Machine Types" in wb.sheetnames:
            ws = wb["Machine Types"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                fee = float(row[1]) if row[1] else 0
                profit = float(row[2]) if len(row) > 2 and row[2] else 0
                existing = await db.machine_types.find_one({"name": name})
                if existing:
                    await db.machine_types.update_one({"name": name}, {"$set": {"monthly_fee": fee, "daily_profit": profit}})
                else:
                    mt = MachineType(name=name, monthly_fee=fee, daily_profit=profit)
                    await db.machine_types.insert_one(mt.model_dump())
                count += 1
            results["machine_types"] = count
        
        # 2. Customers
        if "Customers" in wb.sheetnames:
            ws = wb["Customers"]
            machine_types_list = await db.machine_types.find({}, {"_id": 0}).to_list(100)
            mt_map = {mt["name"].upper(): mt for mt in machine_types_list}
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                phone = str(row[1]).strip() if row[1] else ""
                machines_str = str(row[2]).strip() if row[2] else ""
                cost = float(row[3]) if row[3] else 0
                fee = float(row[4]) if row[4] else 0
                status = str(row[5]).strip().lower() if row[5] else "active"
                prepaid = int(row[6]) if row[6] else 0
                notes = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                
                # Parse machines
                import re
                machines = []
                total_fee = 0
                if machines_str:
                    parts = re.findall(r'(\d+)\s*x\s*([a-zA-Z0-9\-_]+)', machines_str, re.IGNORECASE)
                    for qty_str, mname in parts:
                        qty = int(qty_str)
                        normalized = normalize_machine_name(mname)
                        if normalized in mt_map:
                            mt = mt_map[normalized]
                            machines.append({"machine_type_id": mt["id"], "machine_name": mt["name"], "quantity": qty})
                            total_fee += mt["monthly_fee"] * qty
                
                if total_fee == 0:
                    total_fee = fee
                if status in ['paused', 'pause', 'inactive', 'off']:
                    status = 'paused'
                else:
                    status = 'active'
                
                if mode == "merge":
                    existing = await db.customers.find_one({"name": name, "phone": phone})
                    if existing:
                        await db.customers.update_one({"id": existing["id"]}, {"$set": {"machines": machines, "total_cost": cost, "total_fee": total_fee, "status": status, "prepaid_months": prepaid, "notes": notes}})
                        count += 1
                        continue
                
                customer = Customer(name=name, phone=phone, machines=machines, total_cost=cost, total_fee=total_fee, status=status, prepaid_months=prepaid, notes=notes)
                await db.customers.insert_one(customer.model_dump())
                count += 1
            results["customers"] = count
        
        # 3. Payments
        if "Payments" in wb.sheetnames:
            ws = wb["Payments"]
            count = 0
            # Build customer name -> id map
            all_customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
            cust_name_map = {c["name"]: c["id"] for c in all_customers}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                cust_name = str(row[0]).strip()
                month = str(row[1]).strip() if row[1] else ""
                amount = float(row[2]) if row[2] else 0
                status = str(row[3]).strip().lower() if row[3] else "unpaid"
                paid_at = str(row[4]).strip() if row[4] else None
                
                cust_id = cust_name_map.get(cust_name, "")
                
                if mode == "merge" and cust_id:
                    existing = await db.payments.find_one({"customer_id": cust_id, "month": month})
                    if existing:
                        await db.payments.update_one({"id": existing["id"]}, {"$set": {"amount": amount, "status": status, "paid_at": paid_at}})
                        count += 1
                        continue
                
                payment = Payment(customer_id=cust_id, customer_name=cust_name, month=month, amount=amount, status=status, paid_at=paid_at if paid_at and paid_at != "None" else None)
                await db.payments.insert_one(payment.model_dump())
                count += 1
            results["payments"] = count
        
        # 4. Customer Accounts
        if "Customer Accounts" in wb.sheetnames:
            ws = wb["Customer Accounts"]
            all_customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
            cust_name_map = {c["name"].lower(): c["id"] for c in all_customers}
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                username = str(row[0]).strip().lower()
                password = str(row[1]).strip() if row[1] else ""
                worker_name = str(row[2]).strip() if row[2] else ""
                api_key = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                watcher_key = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                
                # Find customer by username/worker_name
                cust_id = cust_name_map.get(username, cust_name_map.get(worker_name.lower(), ""))
                
                if mode == "merge":
                    existing = await db.customer_accounts.find_one({"username": username})
                    if existing:
                        update = {"password": password, "worker_name": worker_name}
                        if api_key:
                            update["viabtc_api_key"] = api_key
                        if watcher_key:
                            update["watcher_key"] = watcher_key
                        await db.customer_accounts.update_one({"id": existing["id"]}, {"$set": update})
                        count += 1
                        continue
                
                account = CustomerAccount(customer_id=cust_id, username=username, password=password, worker_name=worker_name, viabtc_api_key=api_key)
                doc = account.model_dump()
                if watcher_key:
                    doc["watcher_key"] = watcher_key
                await db.customer_accounts.insert_one(doc)
                count += 1
            results["customer_accounts"] = count
        
        # 5. Maintenance Logs
        if "Maintenance Logs" in wb.sheetnames:
            ws = wb["Maintenance Logs"]
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[2]:
                    continue
                log = MaintenanceLog(customer_id=str(row[0] or ""), machine_info=str(row[1] or ""), description=str(row[2] or ""))
                if row[3]:
                    log.date = str(row[3])
                await db.maintenance_logs.insert_one(log.model_dump())
                count += 1
            results["maintenance_logs"] = count
        
        # 6. Settings
        if "Settings" in wb.sheetnames:
            ws = wb["Settings"]
            update_data = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] and row[1]:
                    update_data[str(row[0])] = str(row[1])
            if update_data:
                await db.settings.update_one({"id": "settings"}, {"$set": update_data}, upsert=True)
                results["settings"] = len(update_data)
        
        # 7. Farm Stats
        if "Farm Stats" in wb.sheetnames:
            ws = wb["Farm Stats"]
            update_data = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] and row[1]:
                    key = str(row[0])
                    val = str(row[1])
                    try:
                        update_data[key] = int(val)
                    except ValueError:
                        update_data[key] = val
            if update_data:
                await db.farm_stats.update_one({"id": "farm_stats"}, {"$set": update_data}, upsert=True)
                results["farm_stats"] = len(update_data)
        
    except Exception as e:
        return {"success": False, "error": str(e), "results": results}
    
    return {"success": True, "message": "Backup imported successfully!", "results": results}


# ==================== INIT DEFAULT MACHINE TYPES ====================

@api_router.post("/init")
async def init_default_data():
    """Initialize default machine types"""
    default_machines = [
        {"name": "L1", "monthly_fee": 100},
        {"name": "L7", "monthly_fee": 300},
        {"name": "L9", "monthly_fee": 250},
        {"name": "Ks5pro", "monthly_fee": 250},
        {"name": "Z15pro", "monthly_fee": 250},
    ]
    
    created = 0
    for machine in default_machines:
        existing = await db.machine_types.find_one({"name": machine["name"]})
        if not existing:
            mt = MachineType(**machine)
            await db.machine_types.insert_one(mt.model_dump())
            created += 1
    
    # Initialize settings
    settings = await db.settings.find_one({"id": "settings"})
    if not settings:
        await get_settings()
    
    return {"message": f"Initialized {created} machine types"}

@api_router.get("/")
async def root():
    return {"message": "WKBeast Crypto Farm Manager API"}

# ==================== CUSTOMER PORTAL ====================

@api_router.post("/portal/login")
async def customer_login(username: str, password: str):
    """Customer login - supports multiple customers with same username"""
    # Find ALL accounts with this username
    accounts = await db.customer_accounts.find({
        "username": username.lower().strip(),
        "password": password
    }, {"_id": 0}).to_list(100)
    
    if not accounts:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    # Get ALL customers linked to these accounts
    customer_ids = [acc["customer_id"] for acc in accounts]
    customers = await db.customers.find({"id": {"$in": customer_ids}}, {"_id": 0}).to_list(100)
    
    if not customers:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Use the first account's API key for all (they share the same worker)
    primary_account = accounts[0]
    
    # Merge customer data - combine machines, calculate total fee
    merged_customer = {
        "id": customer_ids[0],  # Primary ID
        "all_customer_ids": customer_ids,  # Store all IDs for dashboard
        "name": customers[0].get("name", ""),
        "phone": customers[0].get("phone", ""),
        "status": "active" if any(c.get("status") == "active" for c in customers) else "paused",
        "machines": [],
        "total_fee": 0
    }
    
    # Merge machines and fees from all customers
    for c in customers:
        merged_customer["machines"].extend(c.get("machines", []))
        merged_customer["total_fee"] += c.get("total_fee", 0)
    
    return {
        "success": True,
        "account": primary_account,
        "customer": merged_customer
    }

@api_router.get("/portal/dashboard/{customer_id}")
async def get_customer_dashboard(customer_id: str, all_ids: str = ""):
    """Get customer dashboard data - supports merged accounts"""
    
    # Parse all customer IDs (comma-separated) or use single ID
    customer_ids = [cid.strip() for cid in all_ids.split(",") if cid.strip()] if all_ids else [customer_id]
    
    # Fetch all customers
    customers = await db.customers.find({"id": {"$in": customer_ids}}, {"_id": 0}).to_list(100)
    if not customers:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Merge customer data
    merged_customer = {
        "id": customer_ids[0],
        "name": customers[0].get("name", ""),
        "phone": customers[0].get("phone", ""),
        "status": "active" if any(c.get("status") == "active" for c in customers) else "paused",
        "machines": [],
        "total_fee": 0
    }
    
    for c in customers:
        merged_customer["machines"].extend(c.get("machines", []))
        merged_customer["total_fee"] += c.get("total_fee", 0)
    
    # Get machine types with daily profit
    machine_types = await db.machine_types.find({}, {"_id": 0}).to_list(100)
    machine_types_map = {mt["id"]: mt for mt in machine_types}
    
    # Enrich customer machines with daily profit info
    enriched_machines = []
    total_daily_profit = 0
    for m in merged_customer.get("machines", []):
        mt = machine_types_map.get(m["machine_type_id"], {})
        daily_profit = mt.get("daily_profit", 0) * m.get("quantity", 1)
        total_daily_profit += daily_profit
        enriched_machines.append({
            **m,
            "machine_type_name": mt.get("name", m.get("machine_name", "Unknown")),
            "daily_profit": mt.get("daily_profit", 0),
            "total_daily_profit": daily_profit
        })
    
    # Get machine statuses from ALL customers
    machine_statuses = await db.machine_statuses.find({"customer_id": {"$in": customer_ids}}, {"_id": 0}).to_list(100)
    
    # Get payments from ALL customers
    payments = await db.payments.find({"customer_id": {"$in": customer_ids}}, {"_id": 0}).to_list(100)
    
    # Get maintenance logs from ALL customers
    logs = await db.maintenance_logs.find({"customer_id": {"$in": customer_ids}}, {"_id": 0}).to_list(100)
    
    # Get repairs for this customer
    repairs = await db.repairs.find({"customer_id": {"$in": customer_ids}}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Get farm stats with fluctuation
    import random
    farm_stats = await db.farm_stats.find_one({"id": "farm_stats"}, {"_id": 0})
    if not farm_stats:
        farm_stats = {"machines_online": 2430, "machines_offline": 10, "total_hashrate": "850 TH/s", "fluctuation": 5}
    
    fluct = farm_stats.get("fluctuation", 5)
    farm_stats["machines_online_display"] = farm_stats["machines_online"] + random.randint(-fluct, fluct)
    farm_stats["machines_offline_display"] = max(0, farm_stats["machines_offline"] + random.randint(-2, 2))
    
    return {
        "customer": merged_customer,
        "enriched_machines": enriched_machines,
        "total_daily_profit": total_daily_profit,
        "total_monthly_profit": total_daily_profit * 30,
        "machine_statuses": machine_statuses,
        "payments": payments,
        "repairs": repairs,
        "maintenance_logs": logs,
        "farm_stats": farm_stats
    }

@api_router.get("/portal/crypto-prices")
async def get_crypto_prices():
    """Get live crypto prices for earnings calculation"""
    import aiohttp
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                "https://api.coingecko.com/api/v3/simple/price?ids=litecoin,kaspa,zcash&vs_currencies=usd",
                timeout=10
            ) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    return {
                        "ltc": data.get("litecoin", {}).get("usd", 0),
                        "kas": data.get("kaspa", {}).get("usd", 0),
                        "zec": data.get("zcash", {}).get("usd", 0)
                    }
    except Exception as e:
        pass
    
    # Fallback prices
    return {"ltc": 85, "kas": 0.12, "zec": 35}

# ==================== CUSTOMER ACCOUNTS (Admin) ====================

@api_router.get("/customer-accounts")
async def get_customer_accounts():
    accounts = await db.customer_accounts.find({}, {"_id": 0}).to_list(1000)
    return accounts

@api_router.post("/customer-accounts")
async def create_customer_account(data: CustomerAccountCreate):
    # Check if account already exists
    existing = await db.customer_accounts.find_one({"customer_id": data.customer_id})
    if existing:
        raise HTTPException(status_code=400, detail="Account already exists for this customer")
    
    account = CustomerAccount(
        customer_id=data.customer_id,
        username=data.username.lower().strip().replace(" ", ""),
        password=data.password,
        worker_name=data.worker_name
    )
    await db.customer_accounts.insert_one(account.model_dump())
    return account

@api_router.put("/customer-accounts/{account_id}")
async def update_customer_account(account_id: str, data: dict):
    if "username" in data:
        data["username"] = data["username"].lower().strip().replace(" ", "")
    
    result = await db.customer_accounts.find_one_and_update(
        {"id": account_id},
        {"$set": data},
        return_document=True,
        projection={"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Account not found")
    return result

@api_router.delete("/customer-accounts/{account_id}")
async def delete_customer_account(account_id: str):
    result = await db.customer_accounts.delete_one({"id": account_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Account not found")
    return {"message": "Deleted"}

@api_router.put("/customer-accounts/{account_id}/watcher")
async def update_watcher_key(account_id: str, watcher_url: str):
    """Save watcher link for a customer account. Extract access_key from URL."""
    import re
    
    # Extract access_key from watcher URL
    # URL format: https://www.viabtc.com/en/observer/worker?access_key=xxx&coin=LTC
    match = re.search(r'access_key=([a-f0-9]+)', watcher_url)
    if not match:
        raise HTTPException(status_code=400, detail="Invalid watcher URL. Could not find access_key.")
    
    watcher_key = match.group(1)
    
    result = await db.customer_accounts.find_one_and_update(
        {"id": account_id},
        {"$set": {"watcher_key": watcher_key}},
        return_document=True,
        projection={"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Account not found")
    return {"success": True, "watcher_key": watcher_key}

# ==================== MAINTENANCE LOGS ====================

@api_router.get("/maintenance-logs")
async def get_maintenance_logs(customer_id: Optional[str] = None):
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    logs = await db.maintenance_logs.find(query, {"_id": 0}).to_list(1000)
    return logs

@api_router.post("/maintenance-logs")
async def create_maintenance_log(data: MaintenanceLogCreate):
    log = MaintenanceLog(
        customer_id=data.customer_id,
        machine_info=data.machine_info,
        description=data.description
    )
    await db.maintenance_logs.insert_one(log.model_dump())
    return log

@api_router.delete("/maintenance-logs/{log_id}")
async def delete_maintenance_log(log_id: str):
    result = await db.maintenance_logs.delete_one({"id": log_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Log not found")
    return {"message": "Deleted"}

# ==================== FARM STATS (Admin) ====================

@api_router.get("/farm-stats")
async def get_farm_stats():
    stats = await db.farm_stats.find_one({"id": "farm_stats"}, {"_id": 0})
    if not stats:
        new_stats = FarmStats()
        doc = new_stats.model_dump()
        await db.farm_stats.insert_one(doc)
        stats = await db.farm_stats.find_one({"id": "farm_stats"}, {"_id": 0})
    return stats

@api_router.put("/farm-stats")
async def update_farm_stats(data: FarmStatsUpdate):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    result = await db.farm_stats.find_one_and_update(
        {"id": "farm_stats"},
        {"$set": update_data},
        upsert=True,
        return_document=True,
        projection={"_id": 0}
    )
    return result

# ==================== MACHINE STATUS (Admin) ====================

@api_router.get("/machine-statuses")
async def get_machine_statuses(customer_id: Optional[str] = None):
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    statuses = await db.machine_statuses.find(query, {"_id": 0}).to_list(1000)
    return statuses

@api_router.post("/machine-statuses")
async def create_or_update_machine_status(customer_id: str, worker_name: str, status: str = "online", hashrate: str = "0 TH/s", temperature: str = "0°C", uptime: str = "0h"):
    existing = await db.machine_statuses.find_one({"customer_id": customer_id, "worker_name": worker_name})
    
    if existing:
        result = await db.machine_statuses.find_one_and_update(
            {"customer_id": customer_id, "worker_name": worker_name},
            {"$set": {
                "status": status,
                "hashrate": hashrate,
                "temperature": temperature,
                "uptime": uptime,
                "last_updated": datetime.now(timezone.utc).isoformat()
            }},
            return_document=True,
            projection={"_id": 0}
        )
        return result
    else:
        machine_status = MachineStatus(
            customer_id=customer_id,
            worker_name=worker_name,
            status=status,
            hashrate=hashrate,
            temperature=temperature,
            uptime=uptime
        )
        await db.machine_statuses.insert_one(machine_status.model_dump())
        return machine_status

@api_router.delete("/machine-statuses/{status_id}")
async def delete_machine_status(status_id: str):
    result = await db.machine_statuses.delete_one({"id": status_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Status not found")
    return {"message": "Deleted"}

# ==================== VIABTC SETTINGS ====================

@api_router.get("/viabtc-settings")
async def get_viabtc_settings():
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    if not settings:
        settings = ViaBTCSettings().model_dump()
    # Don't return secret key in full
    if settings.get("secret_key"):
        settings["secret_key_masked"] = "****" + settings["secret_key"][-4:] if len(settings["secret_key"]) > 4 else "****"
    return settings

@api_router.put("/viabtc-settings")
async def update_viabtc_settings(access_key: str = "", secret_key: str = "", enabled: bool = False, watcher_key: str = ""):
    """Update ViaBTC settings including optional watcher key for main account"""
    update_data = {
        "access_key": access_key,
        "secret_key": secret_key,
        "enabled": enabled
    }
    if watcher_key:
        # Extract access_key from watcher URL if full URL provided
        import re
        match = re.search(r'access_key=([a-f0-9]+)', watcher_key)
        update_data["watcher_key"] = match.group(1) if match else watcher_key
    
    result = await db.viabtc_settings.find_one_and_update(
        {"id": "viabtc_settings"},
        {"$set": update_data},
        upsert=True,
        return_document=True,
        projection={"_id": 0}
    )
    return result

@api_router.put("/viabtc-watcher")
async def update_main_watcher_key(watcher_url: str):
    """Save watcher link for main account"""
    import re
    
    # Extract access_key from watcher URL
    match = re.search(r'access_key=([a-f0-9]+)', watcher_url)
    if not match:
        raise HTTPException(status_code=400, detail="Invalid watcher URL. Could not find access_key.")
    
    watcher_key = match.group(1)
    
    result = await db.viabtc_settings.find_one_and_update(
        {"id": "viabtc_settings"},
        {"$set": {"watcher_key": watcher_key}},
        upsert=True,
        return_document=True,
        projection={"_id": 0}
    )
    return {"success": True, "watcher_key": watcher_key}

@api_router.post("/viabtc-test")
async def test_viabtc_connection():
    """Test ViaBTC connection using watcher key (primary) or API key (fallback)"""
    import aiohttp
    
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    
    # Always fetch server IP for display
    server_ip = "unknown"
    try:
        async with aiohttp.ClientSession() as ip_session:
            async with ip_session.get("https://api.ipify.org", timeout=5) as ip_resp:
                server_ip = (await ip_resp.text()).strip()
    except:
        pass
    
    # Try watcher key first (bypasses Cloudflare)
    watcher_key = settings.get("watcher_key") if settings else None
    if watcher_key:
        try:
            async with aiohttp.ClientSession() as session:
                url = f"https://www.viabtc.com/res/observer/worker?access_key={watcher_key}&coin=LTC"
                async with session.get(url, timeout=15) as resp:
                    data = await resp.json()
                    if resp.status == 200 and data.get("code") == 0:
                        workers = data.get("data", {}).get("data", [])
                        active = data.get("data", {}).get("active", 0)
                        return {
                            "success": True,
                            "message": f"Watcher connection successful! Found {len(workers)} workers ({active} active).",
                            "mode": "watcher",
                            "data": {"workers": len(workers), "active": active},
                            "server_ip": server_ip
                        }
                    else:
                        return {
                            "success": False,
                            "error": data.get("message", "Unknown error"),
                            "message": f"Watcher key error: {data.get('message', 'Unknown')}",
                            "mode": "watcher",
                            "server_ip": server_ip
                        }
        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "message": f"Watcher connection failed: {str(e)}",
                "mode": "watcher",
                "server_ip": server_ip
            }
    
    # No watcher key - check if API keys exist
    if not settings or not settings.get("access_key"):
        return {
            "success": False,
            "error": "No watcher key or API keys configured",
            "message": "Add a Watcher Link from ViaBTC → Observer → Copy Link",
            "server_ip": server_ip
        }
    
    try:
        # ViaBTC API - get hashrate (public endpoint with API key)
        tonce = str(int(time.time() * 1000))
        
        # Create params and query string
        params = {"coin": "LTC", "tonce": tonce}
        query_string = urlencode(params)
        
        # Generate signature using HMAC-SHA256
        signature = hmac.new(
            secret_key.encode('utf-8'),
            query_string.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        # Headers as per ViaBTC API docs
        headers = {
            "X-API-KEY": access_key,
            "X-SIGNATURE": signature
        }
        
        url = f"https://www.viabtc.com/res/openapi/v1/hashrate?{query_string}"
        
        # Get server IP first
        try:
            async with aiohttp.ClientSession() as ip_session:
                async with ip_session.get("https://api.ipify.org", timeout=5, proxy=_proxy()) as ip_resp:
                    server_ip = (await ip_resp.text()).strip()
        except:
            server_ip = "unknown"
        
        # Increase timeout and add retry
        timeout = aiohttp.ClientTimeout(total=30, connect=15)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url, headers=headers, proxy=_proxy()) as resp:
                data = await resp.json()
                
                if resp.status == 200 and data.get("code") == 0:
                    return {
                        "success": True,
                        "message": "Connection successful! API is working.",
                        "data": data.get("data", {}),
                        "server_ip": server_ip
                    }
                else:
                    error_msg = data.get("message", "Unknown error")
                    error_code = data.get("code")
                    
                    return {
                        "success": False,
                        "error": error_msg,
                        "message": f"API error: {error_msg}" if error_code != 12004 else f"IP not whitelisted. Add server IP: {server_ip}",
                        "code": error_code,
                        "server_ip": server_ip
                    }
    except aiohttp.ClientError as e:
        # Get server IP for error message
        try:
            async with aiohttp.ClientSession() as ip_session:
                async with ip_session.get("https://api.ipify.org", timeout=5, proxy=_proxy()) as ip_resp:
                    server_ip = (await ip_resp.text()).strip()
        except:
            server_ip = "unknown"
        
        return {
            "success": False,
            "error": str(e),
            "message": f"Network error - could not connect to ViaBTC",
            "server_ip": server_ip
        }
    except Exception as e:
        # Get server IP for error message
        try:
            async with aiohttp.ClientSession() as ip_session:
                async with ip_session.get("https://api.ipify.org", timeout=5, proxy=_proxy()) as ip_resp:
                    server_ip = (await ip_resp.text()).strip()
        except:
            server_ip = "unknown"
        
        return {
            "success": False,
            "error": str(e),
            "message": f"Error: {str(e)}",
            "server_ip": server_ip
        }

@api_router.get("/viabtc/earnings/{coin}")
async def get_viabtc_earnings(coin: str = "LTC"):
    """Fetch earnings/profit data from ViaBTC API"""
    import aiohttp
    import hashlib
    import hmac
    import time
    from urllib.parse import urlencode
    
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    if not settings or not settings.get("enabled"):
        return {"success": False, "error": "ViaBTC integration not enabled"}
    
    access_key = settings.get("access_key", "")
    secret_key = settings.get("secret_key", "")
    
    if not access_key or not secret_key:
        return {"success": False, "error": "API keys not configured"}
    
    try:
        tonce = str(int(time.time() * 1000))
        
        # Fetch profit data
        params = {"coin": coin.upper(), "tonce": tonce}
        query_string = urlencode(params)
        
        signature = hmac.new(
            secret_key.encode('utf-8'),
            query_string.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        headers = {
            "X-API-KEY": access_key,
            "X-SIGNATURE": signature
        }
        
        url = f"https://www.viabtc.com/res/openapi/v1/profit?{query_string}"
        
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers, timeout=15, proxy=_proxy()) as resp:
                data = await resp.json()
                
                if resp.status == 200 and data.get("code") == 0:
                    return {
                        "success": True,
                        "coin": coin.upper(),
                        "data": data.get("data", {})
                    }
                else:
                    return {
                        "success": False,
                        "error": data.get("message", "Unknown error"),
                        "code": data.get("code")
                    }
                    
    except Exception as e:
        return {"success": False, "error": str(e)}

@api_router.get("/viabtc/customer-earnings/{account_id}")
async def get_customer_earnings(account_id: str):
    """Fetch earnings for a specific customer using their API key"""
    import aiohttp
    import hashlib
    import hmac
    import time
    from urllib.parse import urlencode
    
    # Get customer account
    account = await db.customer_accounts.find_one({"id": account_id}, {"_id": 0})
    if not account:
        return {"success": False, "error": "Account not found"}
    
    access_key = account.get("viabtc_api_key", "")
    secret_key = account.get("viabtc_secret_key", "")
    
    if not access_key or not secret_key:
        # Try main settings for turkbeast
        settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
        if settings:
            access_key = settings.get("access_key", "")
            secret_key = settings.get("secret_key", "")
    
    if not access_key or not secret_key:
        return {"success": False, "error": "No API keys configured"}
    
    try:
        all_earnings = {}
        balances = {}
        
        async with aiohttp.ClientSession() as session:
            # First get account balances (includes ALL coins from merged mining)
            tonce = str(int(time.time() * 1000))
            params = {"tonce": tonce}
            query_string = urlencode(params)
            
            signature = hmac.new(
                secret_key.encode('utf-8'),
                query_string.encode('utf-8'),
                hashlib.sha256
            ).hexdigest()
            
            headers = {
                "X-API-KEY": access_key,
                "X-SIGNATURE": signature
            }
            
            account_url = f"https://www.viabtc.com/res/openapi/v1/account?{query_string}"
            
            try:
                async with session.get(account_url, headers=headers, timeout=10, proxy=_proxy()) as resp:
                    data = await resp.json()
                    if resp.status == 200 and data.get("code") == 0:
                        balance_list = data.get("data", {}).get("balance", [])
                        for b in balance_list:
                            coin = b.get("coin", "")
                            amount = float(b.get("amount", 0) or 0)
                            if coin and amount > 0:
                                balances[coin] = amount
            except:
                pass
            
            # Get detailed profit for main coins
            main_coins = ["LTC", "KAS", "BTC", "ZEC"]
            for coin in main_coins:
                tonce = str(int(time.time() * 1000))
                params = {"coin": coin, "tonce": tonce}
                query_string = urlencode(params)
                
                signature = hmac.new(
                    secret_key.encode('utf-8'),
                    query_string.encode('utf-8'),
                    hashlib.sha256
                ).hexdigest()
                
                headers = {
                    "X-API-KEY": access_key,
                    "X-SIGNATURE": signature
                }
                
                url = f"https://www.viabtc.com/res/openapi/v1/profit?{query_string}"
                
                try:
                    async with session.get(url, headers=headers, timeout=10, proxy=_proxy()) as resp:
                        data = await resp.json()
                        
                        if resp.status == 200 and data.get("code") == 0:
                            profit_data = data.get("data", {})
                            total = float(profit_data.get("total_profit", 0) or 0)
                            if total > 0:
                                all_earnings[coin] = {
                                    "total_profit": profit_data.get("total_profit", 0),
                                    "pps_profit": profit_data.get("pps_profit", 0),
                                    "pplns_profit": profit_data.get("pplns_profit", 0),
                                    "balance": balances.get(coin, 0)
                                }
                except:
                    pass
            
            # Add merged mining coins from balance (DOGE, BELLS, PEP, etc.)
            merged_coins = ["DOGE", "BELLS", "PEP", "DINGO", "SHIC", "JKC", "LKY"]
            for coin in merged_coins:
                if coin in balances and balances[coin] > 0:
                    all_earnings[coin] = {
                        "total_profit": balances[coin],
                        "balance": balances[coin],
                        "merged_mining": True
                    }
        
        return {
            "success": True,
            "account_id": account_id,
            "earnings": all_earnings,
            "balances": balances
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}

# ==================== VIABTC WORKERS DATA ====================

@api_router.get("/viabtc/subaccounts")
async def get_viabtc_subaccounts():
    """Fetch all sub-accounts from ViaBTC API with their API keys"""
    import aiohttp
    import hashlib
    import hmac
    import time
    from urllib.parse import urlencode
    
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    if not settings or not settings.get("enabled"):
        return {"success": False, "error": "ViaBTC integration not enabled", "subaccounts": []}
    
    if not settings.get("access_key") or not settings.get("secret_key"):
        return {"success": False, "error": "API keys not configured", "subaccounts": []}
    
    access_key = settings["access_key"]
    secret_key = settings["secret_key"]
    
    try:
        tonce = str(int(time.time() * 1000))
        params = {"tonce": tonce, "limit": 100}
        query_string = urlencode(params)
        
        signature = hmac.new(
            secret_key.encode('utf-8'),
            query_string.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        headers = {
            "X-API-KEY": access_key,
            "X-SIGNATURE": signature
        }
        
        url = f"https://www.viabtc.com/res/openapi/v1/account/sub?{query_string}"
        
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers, timeout=15, proxy=_proxy()) as resp:
                data = await resp.json()
                
                if resp.status == 200 and data.get("code") == 0:
                    return {
                        "success": True,
                        "subaccounts": data.get("data", []),
                        "total": data.get("total", 0)
                    }
                else:
                    return {
                        "success": False,
                        "error": data.get("message", "Unknown error"),
                        "subaccounts": []
                    }
    except Exception as e:
        return {"success": False, "error": str(e), "subaccounts": []}

@api_router.post("/viabtc/sync-accounts")
async def sync_viabtc_accounts():
    """Sync ViaBTC sub-accounts with customer accounts - updates username, worker_name, and API keys"""
    # Get main API settings first
    main_settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    main_access_key = main_settings.get("access_key", "") if main_settings else ""
    main_secret_key = main_settings.get("secret_key", "") if main_settings else ""
    
    # Get all sub-accounts from ViaBTC
    subaccounts_res = await get_viabtc_subaccounts()
    if not subaccounts_res.get("success"):
        return subaccounts_res
    
    # Handle nested data structure
    subaccounts_data = subaccounts_res.get("subaccounts", {})
    subaccounts = subaccounts_data.get("data", []) if isinstance(subaccounts_data, dict) else subaccounts_data
    
    # Get all customer accounts
    customer_accounts = await db.customer_accounts.find({}, {"_id": 0}).to_list(1000)
    
    # Create a map of worker_name -> LIST of customer accounts (multiple accounts can share same worker)
    from collections import defaultdict
    account_map = defaultdict(list)
    for acc in customer_accounts:
        worker = acc.get("worker_name", "").lower()
        if worker:
            account_map[worker].append(acc)
    
    updated = []
    not_found = []
    
    # First, check for main account (turkbeast) - give it the main API key
    main_account_name = "turkbeast"
    if main_account_name in account_map and main_access_key:
        # Update ALL accounts with this worker_name
        for acc in account_map[main_account_name]:
            await db.customer_accounts.update_one(
                {"id": acc["id"]},
                {"$set": {
                    "username": main_account_name,
                    "worker_name": main_account_name,
                    "viabtc_api_key": main_access_key,
                    "viabtc_secret_key": main_secret_key
                }}
            )
        updated.append(f"{main_account_name} (main) x{len(account_map[main_account_name])}")
    
    for sub in subaccounts:
        sub_name = sub.get("account", "").lower()
        api_key = sub.get("api_key", "")
        secret_key = sub.get("secret_key", "")
        
        # Skip if this is the main account (already handled above)
        if sub_name == main_account_name:
            continue
        
        if sub_name in account_map:
            # Update ALL customer accounts with this worker_name
            for acc in account_map[sub_name]:
                await db.customer_accounts.update_one(
                    {"id": acc["id"]},
                    {"$set": {
                        "username": sub_name,
                        "worker_name": sub_name,
                        "viabtc_api_key": api_key, 
                        "viabtc_secret_key": secret_key
                    }}
                )
            count = len(account_map[sub_name])
            updated.append(f"{sub_name}" + (f" x{count}" if count > 1 else ""))
        else:
            not_found.append(sub_name)
    
    return {
        "success": True,
        "updated": updated,
        "not_found": not_found,
        "message": f"Synced {len(updated)} accounts. {len(not_found)} sub-accounts not matched to customers."
    }

@api_router.get("/viabtc/workers")
async def get_viabtc_workers(coin: str = "LTC"):
    """Fetch workers data from ViaBTC API"""
    import aiohttp
    import hashlib
    import hmac
    import time
    from urllib.parse import urlencode
    
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    if not settings or not settings.get("enabled"):
        return {"success": False, "error": "ViaBTC integration not enabled", "workers": []}
    
    if not settings.get("access_key") or not settings.get("secret_key"):
        return {"success": False, "error": "API keys not configured", "workers": []}
    
    access_key = settings["access_key"]
    secret_key = settings["secret_key"]
    
    try:
        tonce = str(int(time.time() * 1000))
        params = {"coin": coin, "limit": 100, "tonce": tonce}
        query_string = urlencode(params)
        
        signature = hmac.new(
            secret_key.encode('utf-8'),
            query_string.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        headers = {
            "X-API-KEY": access_key,
            "X-SIGNATURE": signature
        }
        
        # Get workers list using the correct endpoint
        url = f"https://www.viabtc.com/res/openapi/v1/hashrate/worker?{query_string}"
        
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers, timeout=15, proxy=_proxy()) as resp:
                data = await resp.json()
                
                if resp.status == 200 and data.get("code") == 0:
                    workers = data.get("data", {}).get("data", [])
                    return {
                        "success": True,
                        "workers": workers,
                        "total": data.get("data", {}).get("total", len(workers)),
                        "active": sum(1 for w in workers if w.get("worker_status") == "active"),
                        "inactive": sum(1 for w in workers if w.get("worker_status") != "active")
                    }
                else:
                    return {
                        "success": False,
                        "error": data.get("message", "Unknown error"),
                        "workers": []
                    }
    except Exception as e:
        return {"success": False, "error": str(e), "workers": []}

@api_router.get("/viabtc/customer-workers/{customer_account_id}")
async def get_customer_workers(customer_account_id: str):
    """Fetch workers for a specific customer using their own API key - ALL COINS"""
    import aiohttp
    import hashlib
    import hmac
    import time
    from urllib.parse import urlencode
    
    # Get customer account with their API key
    customer_account = await db.customer_accounts.find_one({"id": customer_account_id}, {"_id": 0})
    if not customer_account:
        return {"success": False, "error": "Customer account not found", "workers": []}
    
    customer_api_key = customer_account.get("viabtc_api_key")
    if not customer_api_key:
        return {"success": False, "error": "Customer API key not set", "workers": []}
    
    # Use customer's own secret key if available, otherwise fall back to main
    customer_secret_key = customer_account.get("viabtc_secret_key")
    if not customer_secret_key:
        # Fall back to main settings for secret key
        settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
        if not settings or not settings.get("secret_key"):
            return {"success": False, "error": "Secret key not configured", "workers": []}
        customer_secret_key = settings["secret_key"]
    
    # Fetch workers for ALL coins
    coins = ["LTC", "KAS", "ZEC", "BTC", "DOGE"]
    all_workers = []
    
    async with aiohttp.ClientSession() as session:
        for coin in coins:
            try:
                tonce = str(int(time.time() * 1000))
                params = {"coin": coin, "limit": 100, "tonce": tonce}
                query_string = urlencode(params)
                
                signature = hmac.new(
                    customer_secret_key.encode('utf-8'),
                    query_string.encode('utf-8'),
                    hashlib.sha256
                ).hexdigest()
                
                headers = {
                    "X-API-KEY": customer_api_key,
                    "X-SIGNATURE": signature
                }
                
                url = f"https://www.viabtc.com/res/openapi/v1/hashrate/worker?{query_string}"
                
                async with session.get(url, headers=headers, timeout=15, proxy=_proxy()) as resp:
                    data = await resp.json()
                    
                    if resp.status == 200 and data.get("code") == 0:
                        workers = data.get("data", {}).get("data", [])
                        # Add coin type to each worker
                        for w in workers:
                            w["coin"] = coin
                        all_workers.extend(workers)
            except Exception as e:
                print(f"Error fetching {coin} workers: {e}")
                continue
    
    return {
        "success": True,
        "workers": all_workers,
        "total": len(all_workers),
        "active": sum(1 for w in all_workers if w.get("worker_status") == "active"),
        "inactive": sum(1 for w in all_workers if w.get("worker_status") != "active")
    }

@api_router.get("/viabtc/hashrate")
async def get_viabtc_hashrate(coin: str = "LTC"):
    """Fetch hashrate data from ViaBTC API"""
    import aiohttp
    import hashlib
    import hmac
    import time
    from urllib.parse import urlencode
    
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    if not settings or not settings.get("enabled"):
        return {"success": False, "error": "ViaBTC integration not enabled"}
    
    if not settings.get("access_key") or not settings.get("secret_key"):
        return {"success": False, "error": "API keys not configured"}
    
    access_key = settings["access_key"]
    secret_key = settings["secret_key"]
    
    try:
        tonce = str(int(time.time() * 1000))
        params = {"coin": coin, "tonce": tonce}
        query_string = urlencode(params)
        
        signature = hmac.new(
            secret_key.encode('utf-8'),
            query_string.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        headers = {
            "X-API-KEY": access_key,
            "X-SIGNATURE": signature
        }
        
        url = f"https://www.viabtc.com/res/openapi/v1/hashrate?{query_string}"
        
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers, timeout=15, proxy=_proxy()) as resp:
                data = await resp.json()
                
                if resp.status == 200 and data.get("code") == 0:
                    return {
                        "success": True,
                        "data": data.get("data", {})
                    }
                else:
                    return {
                        "success": False,
                        "error": data.get("message", "Unknown error")
                    }
    except Exception as e:
        return {"success": False, "error": str(e)}

@api_router.get("/portal/worker-status/{worker_name}")
async def get_worker_status(worker_name: str, coin: str = "LTC", api_key: str = None):
    """Get specific worker status from ViaBTC using customer's own API key"""
    import aiohttp
    import hashlib
    import hmac
    import time
    from urllib.parse import urlencode
    
    # Get main account settings for secret key
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    if not settings or not settings.get("enabled"):
        return {"success": False, "error": "ViaBTC not enabled", "worker": None}
    
    if not settings.get("secret_key"):
        return {"success": False, "error": "Secret key not configured", "worker": None}
    
    # Use customer's API key if provided, otherwise fall back to main
    access_key = api_key if api_key else settings.get("access_key")
    if not access_key:
        return {"success": False, "error": "API key not configured", "worker": None}
    
    secret_key = settings["secret_key"]
    
    try:
        tonce = str(int(time.time() * 1000))
        params = {"coin": coin, "limit": 100, "tonce": tonce}
        query_string = urlencode(params)
        
        signature = hmac.new(
            secret_key.encode('utf-8'),
            query_string.encode('utf-8'),
            hashlib.sha256
        ).hexdigest()
        
        headers = {
            "X-API-KEY": access_key,
            "X-SIGNATURE": signature
        }
        
        url = f"https://www.viabtc.com/res/openapi/v1/hashrate/worker?{query_string}"
        
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers, timeout=15, proxy=_proxy()) as resp:
                data = await resp.json()
                
                if resp.status == 200 and data.get("code") == 0:
                    workers = data.get("data", {}).get("data", [])
                    # Find the specific worker (case-insensitive partial match)
                    worker = next((w for w in workers if worker_name.lower() in w.get("worker_name", "").lower()), None)
                    return {
                        "success": True,
                        "worker": worker,
                        "all_workers": workers
                    }
                else:
                    return {
                        "success": False,
                        "error": data.get("message", "Unknown error"),
                        "worker": None
                    }
    except Exception as e:
        return {"success": False, "error": str(e), "worker": None}

# ==================== AUTO-CREATE CUSTOMER ACCOUNTS ====================

@api_router.get("/admin/worker-mapping")
async def get_worker_mapping():
    """Get mapping between database worker_names and ViaBTC sub-accounts"""
    # Get all ViaBTC sub-accounts
    subaccounts_res = await get_viabtc_subaccounts()
    viabtc_names = set()
    if subaccounts_res.get("success"):
        subaccounts_data = subaccounts_res.get("subaccounts", {})
        subaccounts = subaccounts_data.get("data", []) if isinstance(subaccounts_data, dict) else subaccounts_data
        viabtc_names = {s.get("account", "").lower() for s in subaccounts}
    
    # Get all customer accounts
    customer_accounts = await db.customer_accounts.find({}, {"_id": 0}).to_list(1000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    customer_map = {c["id"]: c for c in customers}
    
    mapping = []
    for acc in customer_accounts:
        worker_name = acc.get("worker_name", "").lower()
        customer = customer_map.get(acc.get("customer_id"), {})
        
        mapping.append({
            "account_id": acc.get("id"),
            "customer_name": customer.get("name", "Unknown"),
            "current_worker_name": worker_name,
            "has_api_key": bool(acc.get("viabtc_api_key")),
            "matched_in_viabtc": worker_name in viabtc_names,
            "suggested_matches": [v for v in viabtc_names if worker_name[:4] in v or v[:4] in worker_name] if not worker_name in viabtc_names else []
        })
    
    unmatched = [m for m in mapping if not m["matched_in_viabtc"]]
    matched = [m for m in mapping if m["matched_in_viabtc"]]
    
    return {
        "total_accounts": len(mapping),
        "matched": len(matched),
        "unmatched": len(unmatched),
        "unmatched_details": unmatched,
        "viabtc_subaccounts": sorted(list(viabtc_names))
    }

@api_router.put("/admin/update-worker-name/{account_id}")
async def update_worker_name(account_id: str, new_worker_name: str):
    """Update a customer's worker_name to match ViaBTC"""
    result = await db.customer_accounts.find_one_and_update(
        {"id": account_id},
        {"$set": {"worker_name": new_worker_name.lower().strip()}},
        return_document=True,
        projection={"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Try to sync API keys for this account
    await sync_viabtc_accounts()
    
    return {"success": True, "account": result}

@api_router.post("/admin/bulk-update-worker-names")
async def bulk_update_worker_names(updates: List[Dict]):
    """Bulk update worker names. Format: [{"account_id": "xxx", "new_worker_name": "yyy"}, ...]"""
    updated = []
    for u in updates:
        account_id = u.get("account_id")
        new_name = u.get("new_worker_name", "").lower().strip()
        if account_id and new_name:
            result = await db.customer_accounts.update_one(
                {"id": account_id},
                {"$set": {"worker_name": new_name}}
            )
            if result.modified_count > 0:
                updated.append(new_name)
    
    # Sync API keys
    await sync_viabtc_accounts()
    
    return {"success": True, "updated": updated}


@api_router.post("/auto-create-accounts")
async def auto_create_customer_accounts():
    """Auto-create accounts for all customers without accounts"""
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    existing_accounts = await db.customer_accounts.find({}, {"_id": 0}).to_list(1000)
    existing_customer_ids = {a["customer_id"] for a in existing_accounts}
    
    created = 0
    for customer in customers:
        if customer["id"] not in existing_customer_ids:
            # Generate username from name
            username = customer["name"].lower().strip().replace(" ", "")
            
            # Generate password from last 4 digits of phone
            phone = customer.get("phone", "").replace(" ", "").replace("-", "").replace("+", "")
            password = phone[-4:] if len(phone) >= 4 else "0000"
            
            # Worker name same as username
            worker_name = username
            
            account = CustomerAccount(
                customer_id=customer["id"],
                username=username,
                password=password,
                worker_name=worker_name
            )
            await db.customer_accounts.insert_one(account.model_dump())
            created += 1
    
    return {"message": f"Created {created} accounts"}

# ==================== HIDDEN WORKERS (manual cleanup of stale ViaBTC entries) ====================
# When a customer's worker is renamed on the miner, ViaBTC keeps the old worker name
# as "offline" for ~7 days before auto-purging. These endpoints let the admin manually
# hide those stale entries from the dashboard without waiting.

def _hidden_worker_key(account: str, machine_name: str, coin: str) -> str:
    return f"{(account or '').strip().lower()}||{(machine_name or '').strip().lower()}||{(coin or '').strip().upper()}"


async def _get_hidden_workers_set():
    """Returns a set of canonical hidden-worker keys for fast lookup."""
    docs = await db.hidden_workers.find({}, {"_id": 0}).to_list(10000)
    return set(_hidden_worker_key(d.get("account", ""), d.get("machine_name", ""), d.get("coin", "")) for d in docs)


@api_router.post("/viabtc/hide-worker")
async def hide_worker(data: dict):
    """Hide a stale ViaBTC worker entry from the dashboard.
    Body: {account, machine_name, coin}
    """
    global _watcher_monitor_cache, _machine_monitor_cache
    account = (data.get("account") or "").strip()
    machine_name = (data.get("machine_name") or "").strip()
    coin = (data.get("coin") or "").strip().upper()
    if not machine_name or not coin:
        raise HTTPException(status_code=400, detail="machine_name and coin are required")

    doc = {
        "id": str(uuid.uuid4()),
        "account": account,
        "machine_name": machine_name,
        "coin": coin,
        "key": _hidden_worker_key(account, machine_name, coin),
        "hidden_at": datetime.now(timezone.utc).isoformat()
    }
    await db.hidden_workers.update_one({"key": doc["key"]}, {"$set": doc}, upsert=True)

    # Bust monitor caches so the dashboard reflects the change immediately
    _watcher_monitor_cache["data"] = None
    _watcher_monitor_cache["timestamp"] = 0
    try:
        _machine_monitor_cache.clear()
    except Exception:
        pass

    return {"success": True, "hidden": doc["key"]}


@api_router.post("/viabtc/unhide-worker")
async def unhide_worker(data: dict):
    """Restore a hidden ViaBTC worker entry. Body: {account, machine_name, coin}"""
    global _watcher_monitor_cache, _machine_monitor_cache
    account = (data.get("account") or "").strip()
    machine_name = (data.get("machine_name") or "").strip()
    coin = (data.get("coin") or "").strip().upper()
    key = _hidden_worker_key(account, machine_name, coin)
    res = await db.hidden_workers.delete_one({"key": key})

    _watcher_monitor_cache["data"] = None
    _watcher_monitor_cache["timestamp"] = 0
    try:
        _machine_monitor_cache.clear()
    except Exception:
        pass

    return {"success": True, "removed": res.deleted_count}


@api_router.get("/viabtc/hidden-workers")
async def list_hidden_workers():
    """List all currently hidden workers."""
    docs = await db.hidden_workers.find({}, {"_id": 0}).to_list(10000)
    return {"success": True, "hidden": docs, "count": len(docs)}


# ==================== WATCHER-BASED MACHINE MONITORING ====================
# Cache for watcher monitor
_watcher_monitor_cache = {"data": None, "timestamp": 0}


@api_router.get("/admin/sync-mismatch")
async def get_sync_mismatch(force_refresh: bool = False, mode: str = "api"):
    """Compare ViaBTC pool workers vs PC-sync local machines and report which
    workers are missing on either side. Helps identify miners that the local PC sync
    couldn't reach (offline LAN, wrong subnet, blocked port, etc).

    mode='api' (default) uses the main account API key to see ALL workers in the pool
    (matches what the dashboard shows in API mode). Falls back to watcher if API fails.

    Returns:
      - viabtc_only: workers seen by ViaBTC but with no matching PC-sync entry.
      - pc_only: customer machines in PC sync but no matching pool worker.
    """
    # 1) Pull current ViaBTC worker list using the SAME source the dashboard uses
    viabtc_data = await get_machine_monitor(force_refresh=force_refresh, mode=mode)

    # If API mode failed (e.g. no API keys), auto-fall back to watcher mode
    if (not viabtc_data or not viabtc_data.get("success")) and mode == "api":
        viabtc_data = await get_machine_monitor(force_refresh=force_refresh, mode="watcher")

    if not viabtc_data or not viabtc_data.get("success"):
        return {
            "success": False,
            "error": viabtc_data.get("error", "Could not fetch ViaBTC worker list") if viabtc_data else "No ViaBTC data",
        }

    # 2) Flatten ALL ViaBTC workers (online + offline) into a list of records
    viabtc_workers = []
    for acc in viabtc_data.get("accounts", []):
        account_label = acc.get("account", "")
        acc_worker_name = acc.get("worker_name", "")
        for w in acc.get("online_workers", []) or []:
            viabtc_workers.append({
                "account": account_label,
                "account_worker": acc_worker_name,
                "worker_name": w.get("name", ""),
                "coin": w.get("coin", ""),
                "status": "online",
                "hashrate": w.get("hashrate", 0),
            })
        for w in acc.get("offline_workers", []) or []:
            viabtc_workers.append({
                "account": account_label,
                "account_worker": acc_worker_name,
                "worker_name": w.get("name", ""),
                "coin": w.get("coin", ""),
                "status": w.get("status", "offline"),
                "hashrate": 0,
            })

    # 3) Pull PC-sync machines, filtered to customer-tagged ones (same set the
    #    Live Machine Control Panel shows under "My Customers").
    pc_machines_all = await db.machine_live_data.find({}, {"_id": 0}).to_list(10000)
    accounts = await db.customer_accounts.find({}, {"_id": 0, "worker_name": 1}).to_list(10000)
    customer_workers = set()
    for acc in accounts:
        wn = (acc.get("worker_name") or "").lower().strip()
        if wn:
            customer_workers.add(wn)

    pc_machines = []
    for m in pc_machines_all:
        worker = (m.get("worker_name") or "").lower().strip()
        if not worker:
            continue
        is_customer = False
        for cw in customer_workers:
            if cw in worker or worker in cw:
                is_customer = True
                break
        if is_customer:
            pc_machines.append(m)

    pc_worker_set = set()
    for m in pc_machines:
        wn = (m.get("worker_name") or "").strip().lower()
        if wn:
            pc_worker_set.add(wn)
            # Add EVERY dot-separated segment so any of them can match a ViaBTC name.
            # Skip pure-numeric segments like "001" which are too generic and would cause
            # false positives across different sub-accounts.
            for seg in wn.split("."):
                if seg and not seg.isdigit():
                    pc_worker_set.add(seg)

    # 4) For each ViaBTC ONLINE worker, check if its name is present in PC sync.
    #    We only care about online ones — offline ViaBTC workers are usually stale entries.
    #    Match logic handles all combos:
    #      - exact: "sarywk" == "sarywk"
    #      - head match: viabtc "sarywk" ↔ pc "sarywk.001" (pc starts with viabtc + ".")
    #      - tail match: viabtc "sarywk" ↔ pc "turkbeast.sarywk" (pc ends with "." + viabtc)
    #      - reverse head: viabtc "sarywk.001" ↔ pc "sarywk" (viabtc starts with pc + ".")
    #      - reverse tail: viabtc "turkbeast.sarywk" ↔ pc "sarywk" (viabtc ends with "." + pc)
    def _matches_pc(name: str) -> bool:
        n = (name or "").strip().lower()
        if not n:
            return False
        if n in pc_worker_set:
            return True
        for pwn in pc_worker_set:
            if not pwn:
                continue
            # head match either direction
            if pwn.startswith(n + ".") or n.startswith(pwn + "."):
                return True
            # tail match either direction
            if pwn.endswith("." + n) or n.endswith("." + pwn):
                return True
        return False

    viabtc_only = []
    seen_keys = set()
    for w in viabtc_workers:
        # Only flag ONLINE workers that aren't synced. Offline ones are likely stale.
        if w["status"] != "online":
            continue
        if _matches_pc(w["worker_name"]):
            continue
        key = (w["worker_name"].lower(), w["coin"])
        if key in seen_keys:
            continue
        seen_keys.add(key)
        viabtc_only.append(w)

    # 5) PC-only: customer machines visible locally but NO matching pool entry at all
    viabtc_worker_names = set()
    for w in viabtc_workers:
        n = (w.get("worker_name") or "").strip().lower()
        if n:
            viabtc_worker_names.add(n)
            for seg in n.split("."):
                if seg and not seg.isdigit():
                    viabtc_worker_names.add(seg)

    pc_only = []
    for m in pc_machines:
        wn = (m.get("worker_name") or "").strip().lower()
        if not wn:
            continue
        match = wn in viabtc_worker_names
        if not match:
            for vbtc_name in viabtc_worker_names:
                if not vbtc_name:
                    continue
                if vbtc_name.startswith(wn + ".") or wn.startswith(vbtc_name + "."):
                    match = True
                    break
                if vbtc_name.endswith("." + wn) or wn.endswith("." + vbtc_name):
                    match = True
                    break
        if not match:
            pc_only.append({
                "ip": m.get("ip", ""),
                "worker_name": m.get("worker_name", ""),
                "model": m.get("model", ""),
                "farm": m.get("farm", ""),
                "status": m.get("status", ""),
            })

    return {
        "success": True,
        "mode": viabtc_data.get("mode", mode),
        "viabtc_only": viabtc_only,           # in pool but missing from PC sync
        "pc_only": pc_only,                    # in PC sync but missing from pool
        "viabtc_total": sum(1 for w in viabtc_workers if w["status"] == "online"),
        "pc_total": len(pc_machines),
        "checked_at": time.time(),
    }


# ==================== END SYNC MISMATCH ====================

@api_router.get("/admin/machine-monitor-watcher")
async def get_machine_monitor_watcher(force_refresh: bool = False):
    """Get real-time MACHINE status using watcher links (no IP whitelist needed)"""
    global _watcher_monitor_cache
    import aiohttp
    
    # Check cache first (unless force refresh)
    current_time = time.time()
    if not force_refresh and _watcher_monitor_cache["data"] and (current_time - _watcher_monitor_cache["timestamp"]) < CACHE_TTL:
        return _watcher_monitor_cache["data"]
    
    # Get customer accounts with watcher keys
    customer_accounts = await db.customer_accounts.find({}, {"_id": 0}).to_list(1000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    
    # Get main account watcher key from settings
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    main_watcher_key = settings.get("watcher_key") if settings else None
    
    # Build customer map
    customer_map = {c["id"]: c for c in customers}
    
    # Collect accounts with watcher keys
    accounts_to_fetch = []
    for acc in customer_accounts:
        watcher_key = acc.get("watcher_key")
        if watcher_key:
            customer = customer_map.get(acc.get("customer_id"), {})
            if customer.get("status") != "paused":  # Skip paused accounts
                accounts_to_fetch.append({
                    "watcher_key": watcher_key,
                    "worker_name": acc.get("worker_name", ""),
                    "display_name": customer.get("name", acc.get("worker_name", "")),
                    "account_id": acc.get("id")
                })
    
    # Add main account if it has watcher key
    if main_watcher_key:
        accounts_to_fetch.append({
            "watcher_key": main_watcher_key,
            "worker_name": "turkbeast",
            "display_name": "turkbeast (Main)",
            "account_id": "main"
        })
    
    if not accounts_to_fetch:
        return {
            "success": False,
            "error": "No watcher keys configured",
            "message": "Add watcher links to customer accounts in Admin Panel"
        }
    
    # Fetch worker data from watcher links
    async def fetch_workers_from_watcher(session, watcher_key, coin):
        """Fetch workers from ViaBTC watcher link - uses observer endpoint that bypasses Cloudflare"""
        try:
            url = f"https://www.viabtc.com/res/observer/worker?access_key={watcher_key}&coin={coin}"
            async with session.get(url, timeout=15) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    if data.get("code") == 0:
                        return {"workers": data.get("data", {}).get("data", []), "error": None}
                    else:
                        return {"workers": [], "error": data.get("message", "API error")}
                else:
                    return {"workers": [], "error": f"HTTP {resp.status}"}
        except Exception as e:
            return {"workers": [], "error": str(e)}
    
    # Get machine IPs from live data (pushed by PC sync)
    live_machines_data = await db.machine_live_data.find({}, {"_id": 0, "worker_name": 1, "ip": 1}).to_list(10000)
    worker_to_ip_map = {}
    for lm in live_machines_data:
        wn = lm.get("worker_name", "").lower().strip()
        if wn and lm.get("ip"):
            worker_to_ip_map[wn] = lm["ip"]

    # Hidden workers (manually cleaned-up stale entries)
    hidden_keys = await _get_hidden_workers_set()
    
    account_stats = []
    api_errors = []
    
    ltc_total_online = 0
    ltc_total_offline = 0
    kas_total_online = 0
    kas_total_offline = 0
    all_online_details = []
    all_offline_details = []
    
    async with aiohttp.ClientSession() as session:
        for acc_info in accounts_to_fetch:
            watcher_key = acc_info["watcher_key"]
            display_name = acc_info["display_name"]
            worker_name = acc_info["worker_name"]
            
            # Fetch LTC and KAS workers
            ltc_result = await fetch_workers_from_watcher(session, watcher_key, "LTC")
            kas_result = await fetch_workers_from_watcher(session, watcher_key, "KAS")
            
            # Check for errors
            if ltc_result["error"] and kas_result["error"]:
                api_errors.append({
                    "account": display_name,
                    "worker_name": worker_name,
                    "reason": ltc_result["error"]
                })
                continue
            
            ltc_online = 0
            ltc_offline = 0
            kas_online = 0
            kas_offline = 0
            offline_workers = []
            online_workers = []
            
            # Process LTC workers
            for w in ltc_result.get("workers", []):
                worker_name_viabtc = w.get("worker_name", "")
                hashrate = int(w.get("hashrate_1hour", 0) or 0)
                status = w.get("worker_status", "")

                is_online = status == "active"
                is_truly_offline = status in ["offline", "unactive"]

                # Skip workers that admin manually hid (stale entries after worker rename)
                if is_truly_offline and _hidden_worker_key(display_name, worker_name_viabtc, "LTC") in hidden_keys:
                    continue

                if is_online:
                    ltc_online += 1
                    online_workers.append({"name": worker_name_viabtc, "coin": "LTC", "hashrate": hashrate})
                elif is_truly_offline:
                    ltc_offline += 1
                    offline_workers.append({"name": worker_name_viabtc, "coin": "LTC", "status": status})

            # Process KAS workers
            for w in kas_result.get("workers", []):
                worker_name_viabtc = w.get("worker_name", "")
                hashrate = int(w.get("hashrate_1hour", 0) or 0)
                status = w.get("worker_status", "")

                is_online = status == "active"
                is_truly_offline = status in ["offline", "unactive"]

                if is_truly_offline and _hidden_worker_key(display_name, worker_name_viabtc, "KAS") in hidden_keys:
                    continue

                if is_online:
                    kas_online += 1
                    online_workers.append({"name": worker_name_viabtc, "coin": "KAS", "hashrate": hashrate})
                elif is_truly_offline:
                    kas_offline += 1
                    offline_workers.append({"name": worker_name_viabtc, "coin": "KAS", "status": status})
            
            if ltc_online + ltc_offline + kas_online + kas_offline > 0:
                ltc_hashrate = sum(w["hashrate"] for w in online_workers if w["coin"] == "LTC")
                kas_hashrate = sum(w["hashrate"] for w in online_workers if w["coin"] == "KAS")
                
                account_stats.append({
                    "account": display_name,
                    "worker_name": worker_name,
                    "ltc_online": ltc_online,
                    "ltc_offline": ltc_offline,
                    "kas_online": kas_online,
                    "kas_offline": kas_offline,
                    "ltc_hashrate": ltc_hashrate,
                    "kas_hashrate": kas_hashrate,
                    "total_online": ltc_online + kas_online,
                    "total_offline": ltc_offline + kas_offline,
                    "offline_workers": offline_workers,
                    "online_workers": online_workers
                })
                
                ltc_total_online += ltc_online
                ltc_total_offline += ltc_offline
                kas_total_online += kas_online
                kas_total_offline += kas_offline
                
                if ltc_online > 0:
                    ltc_workers_list = [w for w in online_workers if w["coin"] == "LTC"]
                    all_online_details.append({
                        "worker": display_name,
                        "worker_name": worker_name,
                        "coin": "LTC",
                        "machines": ltc_online,
                        "hashrate": ltc_hashrate,
                        "workers": ltc_workers_list
                    })
                if kas_online > 0:
                    kas_workers_list = [w for w in online_workers if w["coin"] == "KAS"]
                    all_online_details.append({
                        "worker": display_name,
                        "worker_name": worker_name,
                        "coin": "KAS",
                        "machines": kas_online,
                        "hashrate": kas_hashrate,
                        "workers": kas_workers_list
                    })
                
                for ow in offline_workers:
                    all_offline_details.append({
                        "worker": display_name,
                        "worker_name": worker_name,
                        "machine_name": ow["name"],
                        "coin": ow["coin"],
                        "reason": ow.get("status", "offline")
                    })
    
    total_online = ltc_total_online + kas_total_online
    total_offline = ltc_total_offline + kas_total_offline
    
    result = {
        "success": True,
        "mode": "watcher",
        "stats": {
            "ltc": {"online": ltc_total_online, "offline": ltc_total_offline, "total": ltc_total_online + ltc_total_offline},
            "kas": {"online": kas_total_online, "offline": kas_total_offline, "total": kas_total_online + kas_total_offline},
            "total": {"online": total_online, "offline": total_offline, "total": total_online + total_offline}
        },
        "accounts": account_stats,
        "online_details": all_online_details,
        "offline_details": all_offline_details,
        "api_errors": api_errors,
        "cached_at": current_time
    }
    
    # Update cache
    _watcher_monitor_cache["data"] = result
    _watcher_monitor_cache["timestamp"] = current_time
    
    return result

# ==================== REAL-TIME MACHINE MONITORING ====================
@api_router.get("/admin/machine-monitor")
async def get_machine_monitor(force_refresh: bool = False, mode: str = "api"):
    """Get real-time MACHINE status from ViaBTC. mode=api tries API keys first (falls back to watcher), mode=watcher uses watcher links only."""
    global _machine_monitor_cache
    import aiohttp
    import hashlib
    import hmac
    from urllib.parse import urlencode
    
    # Use separate cache keys per mode
    cache_key = f"{mode}"
    current_time = time.time()
    if not force_refresh and _machine_monitor_cache.get(cache_key) and (current_time - _machine_monitor_cache.get(f"{cache_key}_ts", 0)) < CACHE_TTL:
        return _machine_monitor_cache[cache_key]
    
    # Get customer accounts and customers
    customer_accounts = await db.customer_accounts.find({}, {"_id": 0}).to_list(1000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    customer_map = {c["id"]: c for c in customers}
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    
    # ===== API MODE =====
    if mode == "api":
        main_api_key = settings.get("access_key", "") if settings else ""
        main_secret_key = settings.get("secret_key", "") if settings else ""
        
        if not main_api_key or not main_secret_key or not (settings or {}).get("enabled"):
            # No API keys — return error so frontend can fallback to watcher
            return {"success": False, "error": "API keys not configured or ViaBTC not enabled", "mode": "api"}
        
        # Build account map by API key
        account_by_api_key = {}
        for acc in customer_accounts:
            api_key = acc.get("viabtc_api_key")
            if api_key:
                customer = customer_map.get(acc.get("customer_id"), {})
                if customer.get("status") != "paused":
                    account_by_api_key[api_key] = {
                        "worker_name": acc.get("worker_name", ""),
                        "display_name": customer.get("name", acc.get("worker_name", "")),
                        "secret_key": acc.get("viabtc_secret_key", main_secret_key)
                    }
        account_by_api_key[main_api_key] = {
            "worker_name": "turkbeast",
            "display_name": "turkbeast (Main)",
            "secret_key": main_secret_key
        }
        
        async def fetch_api_workers(session, api_key, secret_key, coin):
            try:
                tonce = str(int(time.time() * 1000))
                params = {"coin": coin, "limit": 100, "tonce": tonce}
                query_string = urlencode(params)
                signature = hmac.new(secret_key.encode(), query_string.encode(), hashlib.sha256).hexdigest()
                headers = {"X-API-KEY": api_key, "X-SIGNATURE": signature}
                url = f"https://www.viabtc.com/res/openapi/v1/hashrate/worker?{query_string}"
                async with session.get(url, headers=headers, timeout=15, proxy=_proxy()) as resp:
                    data = await resp.json()
                    if resp.status == 200 and data.get("code") == 0:
                        return {"workers": data.get("data", {}).get("data", []), "error": None}
                    else:
                        return {"workers": [], "error": data.get("message", "API error")}
            except Exception as e:
                return {"workers": [], "error": str(e)}
        
        # Run API-mode fetch for all accounts
        api_account_stats = []
        api_errors_list = []
        api_ltc_on = api_ltc_off = api_kas_on = api_kas_off = 0
        api_online_details = []
        api_offline_details = []

        # Build worker -> IP/model/farm lookup so offline cards can show details
        live_machines_data_api = await db.machine_live_data.find({}, {"_id": 0, "worker_name": 1, "ip": 1, "model": 1, "farm": 1}).to_list(10000)
        api_worker_to_ip_map = {}
        api_worker_to_meta_map = {}
        for lm in live_machines_data_api:
            wn = (lm.get("worker_name") or "").lower().strip()
            ip = lm.get("ip", "")
            if not (wn and ip):
                continue
            meta = {"ip": ip, "model": lm.get("model", ""), "farm": lm.get("farm", "")}
            if wn not in api_worker_to_ip_map:
                api_worker_to_ip_map[wn] = ip
                api_worker_to_meta_map[wn] = meta
            for seg in wn.split("."):
                if seg and not seg.isdigit() and seg not in api_worker_to_ip_map:
                    api_worker_to_ip_map[seg] = ip
                    api_worker_to_meta_map[seg] = meta

        def _api_lookup(name: str, account_worker: str = "") -> dict:
            n = (name or "").strip().lower()
            if not n:
                return {}
            if n in api_worker_to_ip_map:
                return api_worker_to_meta_map.get(n, {"ip": api_worker_to_ip_map[n]})
            for seg in n.split("."):
                if seg and not seg.isdigit() and seg in api_worker_to_ip_map:
                    return api_worker_to_meta_map.get(seg, {"ip": api_worker_to_ip_map[seg]})
            return {}
        
        async with aiohttp.ClientSession() as session:
            async def fetch_api_account(api_key, info):
                try:
                    ltc_r, kas_r = await asyncio.gather(
                        fetch_api_workers(session, api_key, info["secret_key"], "LTC"),
                        fetch_api_workers(session, api_key, info["secret_key"], "KAS")
                    )
                except Exception as e:
                    return {"error": True, "account": info["display_name"], "worker_name": info["worker_name"], "reason": str(e)}
                
                if ltc_r.get("error") and kas_r.get("error"):
                    return {"error": True, "account": info["display_name"], "worker_name": info["worker_name"], "reason": ltc_r["error"]}
                
                lo = ko = loff = koff = 0
                on_w = []
                off_w = []
                for w in ltc_r.get("workers", []):
                    wn = w.get("worker_name", "")
                    hr = int(w.get("hashrate_1hour", 0) or 0)
                    st = w.get("worker_status", "")
                    if st == "active":
                        lo += 1; on_w.append({"name": wn, "coin": "LTC", "hashrate": hr})
                    elif st in ["offline", "unactive"]:
                        loff += 1; off_w.append({"name": wn, "coin": "LTC", "status": st, "last_active": w.get("last_active", 0)})
                for w in kas_r.get("workers", []):
                    wn = w.get("worker_name", "")
                    hr = int(w.get("hashrate_1hour", 0) or 0)
                    st = w.get("worker_status", "")
                    if st == "active":
                        ko += 1; on_w.append({"name": wn, "coin": "KAS", "hashrate": hr})
                    elif st in ["offline", "unactive"]:
                        koff += 1; off_w.append({"name": wn, "coin": "KAS", "status": st, "last_active": w.get("last_active", 0)})
                
                if lo + loff + ko + koff > 0:
                    return {
                        "account": info["display_name"], "worker_name": info["worker_name"],
                        "ltc_online": lo, "ltc_offline": loff, "kas_online": ko, "kas_offline": koff,
                        "ltc_hashrate": sum(w["hashrate"] for w in on_w if w["coin"] == "LTC"),
                        "kas_hashrate": sum(w["hashrate"] for w in on_w if w["coin"] == "KAS"),
                        "total_online": lo + ko, "total_offline": loff + koff,
                        "offline_workers": off_w, "online_workers": on_w
                    }
                return None
            
            tasks = [fetch_api_account(k, v) for k, v in account_by_api_key.items()]
            results = await asyncio.gather(*tasks)
        
        any_success = False
        for r in results:
            if r and not r.get("error"):
                any_success = True
                api_account_stats.append(r)
                api_ltc_on += r["ltc_online"]; api_ltc_off += r["ltc_offline"]
                api_kas_on += r["kas_online"]; api_kas_off += r["kas_offline"]
                dn = r["account"]; wn = r["worker_name"]
                if r["ltc_online"] > 0:
                    api_online_details.append({"worker": dn, "worker_name": wn, "coin": "LTC", "machines": r["ltc_online"], "hashrate": r["ltc_hashrate"], "workers": [w for w in r["online_workers"] if w["coin"] == "LTC"]})
                if r["kas_online"] > 0:
                    api_online_details.append({"worker": dn, "worker_name": wn, "coin": "KAS", "machines": r["kas_online"], "hashrate": r["kas_hashrate"], "workers": [w for w in r["online_workers"] if w["coin"] == "KAS"]})
                for ow in r["offline_workers"]:
                    meta = _api_lookup(ow["name"], wn)
                    api_offline_details.append({"worker": dn, "worker_name": wn, "machine_name": ow["name"], "coin": ow["coin"], "reason": ow.get("status", "offline"), "last_active": ow.get("last_active", 0), "ip": meta.get("ip", ""), "model": meta.get("model", ""), "farm": meta.get("farm", "")})
            elif r and r.get("error"):
                api_errors_list.append({"account": r.get("account"), "worker_name": r.get("worker_name"), "reason": r.get("reason")})
        
        if any_success:
            t_on = api_ltc_on + api_kas_on; t_off = api_ltc_off + api_kas_off
            api_result = {
                "success": True, "mode": "api",
                "stats": {"ltc": {"online": api_ltc_on, "offline": api_ltc_off, "total": api_ltc_on + api_ltc_off}, "kas": {"online": api_kas_on, "offline": api_kas_off, "total": api_kas_on + api_kas_off}, "total": {"online": t_on, "offline": t_off, "total": t_on + t_off}},
                "accounts": api_account_stats, "online_details": api_online_details, "offline_details": api_offline_details, "api_errors": api_errors_list, "cached_at": current_time
            }
            _machine_monitor_cache[cache_key] = api_result
            _machine_monitor_cache[f"{cache_key}_ts"] = current_time
            return api_result
        
        return {"success": False, "error": "API calls failed for all accounts", "mode": "api", "api_errors": api_errors_list}
    
    # ===== WATCHER MODE =====
    main_watcher_key = settings.get("watcher_key") if settings else None
    
    accounts_to_fetch = []
    seen_watcher_keys = set()
    
    for acc in customer_accounts:
        watcher_key = acc.get("watcher_key")
        if watcher_key and watcher_key not in seen_watcher_keys:
            customer = customer_map.get(acc.get("customer_id"), {})
            if customer.get("status") != "paused":
                seen_watcher_keys.add(watcher_key)
                accounts_to_fetch.append({
                    "watcher_key": watcher_key,
                    "worker_name": acc.get("worker_name", ""),
                    "display_name": customer.get("name", acc.get("worker_name", "")),
                })
    
    if main_watcher_key and main_watcher_key not in seen_watcher_keys:
        accounts_to_fetch.append({
            "watcher_key": main_watcher_key,
            "worker_name": "turkbeast",
            "display_name": "turkbeast (Main)",
        })
    
    if not accounts_to_fetch:
        return {
            "success": False,
            "error": "No watcher keys configured. Add watcher links in Admin Panel.",
            "mode": "watcher"
        }
    
    async def fetch_observer_workers(session, watcher_key, coin):
        try:
            url = f"https://www.viabtc.com/res/observer/worker?access_key={watcher_key}&coin={coin}&limit=200"
            async with session.get(url, timeout=15) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    if data.get("code") == 0:
                        return {"workers": data.get("data", {}).get("data", []), "error": None}
                    else:
                        return {"workers": [], "error": data.get("message", "API error")}
                else:
                    return {"workers": [], "error": f"HTTP {resp.status}"}
        except Exception as e:
            return {"workers": [], "error": str(e)}
    
    # Fetch ALL accounts in parallel
    account_stats = []
    api_errors = []
    ltc_total_online = 0
    ltc_total_offline = 0
    kas_total_online = 0
    kas_total_offline = 0
    all_online_details = []
    all_offline_details = []

    # Hidden workers (manually cleaned-up stale entries)
    hidden_keys = await _get_hidden_workers_set()

    # Build worker -> IP map from PC-sync live data so offline cards can show the IP.
    # Also index by every dot-separated segment + farm + model so we can look up
    # by partial name (e.g. ViaBTC reports "101x129" but PC sync has "nazihwk.101x129").
    live_machines_data = await db.machine_live_data.find({}, {"_id": 0, "worker_name": 1, "ip": 1, "model": 1, "farm": 1}).to_list(10000)
    worker_to_ip_map = {}
    worker_to_meta_map = {}  # ip -> {ip, model, farm}

    def _index_lm(key, ip, meta):
        if key and key not in worker_to_ip_map:
            worker_to_ip_map[key] = ip
            worker_to_meta_map[key] = meta

    for lm in live_machines_data:
        wn = (lm.get("worker_name") or "").lower().strip()
        ip = lm.get("ip", "")
        if not (wn and ip):
            continue
        meta = {"ip": ip, "model": lm.get("model", ""), "farm": lm.get("farm", "")}
        _index_lm(wn, ip, meta)
        # Index every dot-separated segment, skip pure numeric (avoid "001" collisions)
        for seg in wn.split("."):
            if seg and not seg.isdigit():
                _index_lm(seg, ip, meta)

    def _lookup_ip_for_worker(viabtc_name: str, account_worker: str = "") -> dict:
        """Return {'ip', 'model', 'farm'} for a ViaBTC worker name using fuzzy match.
        Returns {} if no confident match — better to show no IP than a wrong one.
        """
        n = (viabtc_name or "").strip().lower()
        if not n:
            return {}
        # 1. exact match
        if n in worker_to_ip_map:
            return worker_to_meta_map.get(n, {"ip": worker_to_ip_map[n]})
        # 2. by tail/head segment (skip pure numeric to avoid "001" collisions)
        for seg in n.split("."):
            if seg and not seg.isdigit() and seg in worker_to_ip_map:
                return worker_to_meta_map.get(seg, {"ip": worker_to_ip_map[seg]})
        return {}
    
    async with aiohttp.ClientSession() as session:
        async def fetch_account_data(acc_info):
            watcher_key = acc_info["watcher_key"]
            display_name = acc_info["display_name"]
            worker_name = acc_info["worker_name"]
            
            # Fetch LTC and KAS workers in parallel
            try:
                ltc_result, kas_result = await asyncio.gather(
                    fetch_observer_workers(session, watcher_key, "LTC"),
                    fetch_observer_workers(session, watcher_key, "KAS")
                )
            except Exception as e:
                return {"error": True, "account": display_name, "worker_name": worker_name, "reason": str(e)}
            
            ltc_error = ltc_result.get("error")
            kas_error = kas_result.get("error")
            
            if ltc_error and kas_error:
                return {"error": True, "account": display_name, "worker_name": worker_name, "reason": ltc_error}
            
            ltc_workers = ltc_result.get("workers", [])
            kas_workers = kas_result.get("workers", [])
            
            ltc_online = 0
            ltc_offline = 0
            kas_online = 0
            kas_offline = 0
            offline_workers = []
            online_workers = []
            
            for w in ltc_workers:
                wname = w.get("worker_name", w.get("name", ""))
                hashrate = int(w.get("hashrate_1hour", 0) or 0)
                status = w.get("worker_status", w.get("status", ""))

                if status in ["offline", "unactive"] and _hidden_worker_key(display_name, wname, "LTC") in hidden_keys:
                    continue

                if status == "active":
                    ltc_online += 1
                    online_workers.append({"name": wname, "coin": "LTC", "hashrate": hashrate})
                elif status in ["offline", "unactive"]:
                    ltc_offline += 1
                    offline_workers.append({"name": wname, "coin": "LTC", "status": status, "last_active": w.get("last_active", 0)})

            for w in kas_workers:
                wname = w.get("worker_name", w.get("name", ""))
                hashrate = int(w.get("hashrate_1hour", 0) or 0)
                status = w.get("worker_status", w.get("status", ""))

                if status in ["offline", "unactive"] and _hidden_worker_key(display_name, wname, "KAS") in hidden_keys:
                    continue

                if status == "active":
                    kas_online += 1
                    online_workers.append({"name": wname, "coin": "KAS", "hashrate": hashrate})
                elif status in ["offline", "unactive"]:
                    kas_offline += 1
                    offline_workers.append({"name": wname, "coin": "KAS", "status": status, "last_active": w.get("last_active", 0)})
            
            if ltc_online + ltc_offline + kas_online + kas_offline > 0:
                ltc_hashrate = sum(w["hashrate"] for w in online_workers if w["coin"] == "LTC")
                kas_hashrate = sum(w["hashrate"] for w in online_workers if w["coin"] == "KAS")
                
                return {
                    "account": display_name,
                    "worker_name": worker_name,
                    "ltc_online": ltc_online,
                    "ltc_offline": ltc_offline,
                    "kas_online": kas_online,
                    "kas_offline": kas_offline,
                    "ltc_hashrate": ltc_hashrate,
                    "kas_hashrate": kas_hashrate,
                    "total_online": ltc_online + kas_online,
                    "total_offline": ltc_offline + kas_offline,
                    "offline_workers": offline_workers,
                    "online_workers": online_workers
                }
            return None
        
        # Execute in batches of 5
        keys_list = accounts_to_fetch
        results = []
        batch_size = 5
        
        for i in range(0, len(keys_list), batch_size):
            batch = keys_list[i:i + batch_size]
            batch_tasks = [fetch_account_data(acc) for acc in batch]
            batch_results = await asyncio.gather(*batch_tasks)
            results.extend(batch_results)
            if i + batch_size < len(keys_list):
                await asyncio.sleep(0.2)
        
        for result in results:
            if result:
                if result.get("error"):
                    api_errors.append({
                        "account": result.get("account"),
                        "worker_name": result.get("worker_name"),
                        "reason": result.get("reason")
                    })
                    continue
                
                account_stats.append(result)
                ltc_total_online += result["ltc_online"]
                ltc_total_offline += result["ltc_offline"]
                kas_total_online += result["kas_online"]
                kas_total_offline += result["kas_offline"]
                
                display_name = result["account"]
                worker_name = result["worker_name"]
                
                if result["ltc_online"] > 0:
                    ltc_workers_list = [w for w in result.get("online_workers", []) if w["coin"] == "LTC"]
                    all_online_details.append({
                        "worker": display_name, "worker_name": worker_name,
                        "coin": "LTC", "machines": result["ltc_online"],
                        "hashrate": result.get("ltc_hashrate", 0), "workers": ltc_workers_list
                    })
                if result["kas_online"] > 0:
                    kas_workers_list = [w for w in result.get("online_workers", []) if w["coin"] == "KAS"]
                    all_online_details.append({
                        "worker": display_name, "worker_name": worker_name,
                        "coin": "KAS", "machines": result["kas_online"],
                        "hashrate": result.get("kas_hashrate", 0), "workers": kas_workers_list
                    })
                
                for ow in result["offline_workers"]:
                    meta = _lookup_ip_for_worker(ow["name"], worker_name)
                    all_offline_details.append({
                        "worker": display_name, "worker_name": worker_name,
                        "machine_name": ow["name"], "coin": ow["coin"],
                        "reason": ow.get("status", "offline"),
                        "last_active": ow.get("last_active", 0),
                        "ip": meta.get("ip", ""),
                        "model": meta.get("model", ""),
                        "farm": meta.get("farm", "")
                    })
    
    total_online = ltc_total_online + kas_total_online
    total_offline = ltc_total_offline + kas_total_offline
    
    result = {
        "success": True,
        "mode": "watcher",
        "stats": {
            "ltc": {"online": ltc_total_online, "offline": ltc_total_offline, "total": ltc_total_online + ltc_total_offline},
            "kas": {"online": kas_total_online, "offline": kas_total_offline, "total": kas_total_online + kas_total_offline},
            "total": {"online": total_online, "offline": total_offline, "total": total_online + total_offline}
        },
        "accounts": account_stats,
        "online_details": all_online_details,
        "offline_details": all_offline_details,
        "api_errors": api_errors,
        "cached_at": current_time
    }
    
    _machine_monitor_cache[cache_key] = result
    _machine_monitor_cache[f"{cache_key}_ts"] = current_time
    
    return result

# ==================== WHATSAPP BOT (Twilio) ====================

from twilio.rest import Client as TwilioClient

def get_twilio_client():
    sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
    token = os.environ.get("TWILIO_AUTH_TOKEN", "")
    if not sid or not token:
        return None
    return TwilioClient(sid, token)

def get_whatsapp_from():
    num = os.environ.get("TWILIO_WHATSAPP_NUMBER", "")
    return f"whatsapp:{num}" if num else ""

def format_phone_whatsapp(phone):
    """Format phone number for WhatsApp. Only adds +961 if it's a Lebanese local number."""
    phone = phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if phone.startswith("+"):
        return phone
    if phone.startswith("00"):
        return "+" + phone[2:]
    # Lebanese local: starts with 0 and is 8 digits (0X XXXXXX) 
    if phone.startswith("0") and len(phone) == 8:
        return "+961" + phone[1:]
    # 7 digits without 0 = Lebanese local
    if len(phone) == 7:
        return "+961" + phone
    # Everything else: already has country code, just add +
    return "+" + phone

@api_router.post("/whatsapp/send")
async def send_whatsapp_message(to: str, message: str):
    """Send a WhatsApp message to a customer"""
    client = get_twilio_client()
    if not client:
        return {"success": False, "error": "Twilio not configured"}
    
    phone = format_phone_whatsapp(to)
    
    try:
        msg = client.messages.create(
            body=message,
            from_=get_whatsapp_from(),
            to=f"whatsapp:{phone}"
        )
        return {"success": True, "sid": msg.sid, "status": msg.status}
    except Exception as e:
        return {"success": False, "error": str(e)}

@api_router.post("/whatsapp/send-payment-reminders")
async def send_payment_reminders(month: str):
    """Send payment reminders to all unpaid customers for a given month via WhatsApp"""
    client = get_twilio_client()
    if not client:
        return {"success": False, "error": "Twilio not configured"}
    
    # Get settings for message template
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    if not settings:
        return {"success": False, "error": "Settings not configured"}
    
    # Get unpaid payments for the month
    payments = await db.payments.find({"month": month, "status": "unpaid"}, {"_id": 0}).to_list(10000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    customer_map = {c["id"]: c for c in customers}
    
    sent = []
    failed = []
    skipped = []
    
    for payment in payments:
        customer = customer_map.get(payment.get("customer_id"))
        if not customer:
            skipped.append({"name": payment.get("customer_name", "Unknown"), "reason": "Customer not found"})
            continue
        
        if customer.get("status") == "paused":
            skipped.append({"name": customer["name"], "reason": "Paused"})
            continue
        
        if not customer.get("whatsapp_enabled", True):
            skipped.append({"name": customer["name"], "reason": "WhatsApp disabled"})
            continue
        phone = customer.get("phone", "")
        if not phone:
            skipped.append({"name": customer["name"], "reason": "No phone number"})
            continue
        
        # Format phone
        phone_clean = format_phone_whatsapp(phone)
        
        # Format message from template
        import urllib.parse
        message = settings.get("message_template", "Payment reminder: ${amount} for {month}").format(
            month=month,
            amount=payment.get("amount", 0),
            whish=settings.get("whish_number", ""),
            usdt=settings.get("usdt_address", ""),
            team=settings.get("team_name", "WKBeast")
        )
        
        try:
            msg = client.messages.create(
                body=message,
                from_=get_whatsapp_from(),
                to=f"whatsapp:{phone_clean}"
            )
            sent.append({"name": customer["name"], "phone": phone, "sid": msg.sid})
        except Exception as e:
            failed.append({"name": customer["name"], "phone": phone, "error": str(e)})
    
    return {
        "success": True,
        "sent": len(sent),
        "failed": len(failed),
        "skipped": len(skipped),
        "details": {"sent": sent, "failed": failed, "skipped": skipped}
    }

@api_router.post("/whatsapp/send-single-reminder")
async def send_single_reminder(customer_id: str, month: str):
    """Send payment reminder to a single customer"""
    client = get_twilio_client()
    if not client:
        return {"success": False, "error": "Twilio not configured"}
    
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        return {"success": False, "error": "Customer not found"}
    
    payment = await db.payments.find_one({"customer_id": customer_id, "month": month}, {"_id": 0})
    if not payment:
        return {"success": False, "error": "No payment found for this month"}
    
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    
    phone = customer.get("phone", "")
    if not phone:
        return {"success": False, "error": "Customer has no phone number"}
    
    phone_clean = format_phone_whatsapp(phone)
    
    message = settings.get("message_template", "Payment reminder: ${amount} for {month}").format(
        month=month,
        amount=payment.get("amount", 0),
        whish=settings.get("whish_number", ""),
        usdt=settings.get("usdt_address", ""),
        team=settings.get("team_name", "WKBeast")
    )
    
    try:
        msg = client.messages.create(
            body=message,
            from_=get_whatsapp_from(),
            to=f"whatsapp:{phone_clean}"
        )
        return {"success": True, "sid": msg.sid, "status": msg.status, "to": phone_clean}
    except Exception as e:
        return {"success": False, "error": str(e)}

@api_router.post("/whatsapp/test-send")
async def test_whatsapp_send(to: str):
    """Test WhatsApp connection by sending a test message"""
    client = get_twilio_client()
    if not client:
        return {"success": False, "error": "Twilio credentials not configured. Add TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN to .env"}
    
    phone = to.replace(" ", "").replace("-", "")
    if not phone.startswith("+"):
        phone = "+" + phone
    
    try:
        # Use offline alert template for test (it's approved for business-initiated)
        template_sid = os.environ.get("TWILIO_TEMPLATE_OFFLINE", "")
        if template_sid:
            import json
            msg = client.messages.create(
                content_sid=template_sid,
                content_variables=json.dumps({"1": "0", "2": "All machines online. Bot connected successfully!"}),
                from_=get_whatsapp_from(),
                to=f"whatsapp:{phone}"
            )
        else:
            msg = client.messages.create(
                body="WKBeast Mining Farm Bot connected!",
                from_=get_whatsapp_from(),
                to=f"whatsapp:{phone}"
            )
        return {"success": True, "sid": msg.sid, "status": msg.status, "message": f"Test message sent to {phone}"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@api_router.post("/whatsapp/notify-offline")
async def notify_offline_machines(admin_phone: str = "+9613022005"):
    """Check for offline machines and send WhatsApp alert to admin"""
    client = get_twilio_client()
    if not client:
        return {"success": False, "error": "Twilio not configured"}
    
    # Fetch current machine status
    import aiohttp
    customer_accounts = await db.customer_accounts.find({}, {"_id": 0}).to_list(1000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    customer_map = {c["id"]: c for c in customers}
    
    # Detect machine type from hashrate
    def detect_machine_type(hashrate_gh):
        if 11 <= hashrate_gh <= 17:
            return "L9"
        elif 7 <= hashrate_gh <= 10:
            return "L7"
        elif 3 <= hashrate_gh <= 6:
            return "L1"
        else:
            return "Unknown"
    
    # Build machine info per customer from DB
    customer_machine_info = {}
    for c in customers:
        machines = c.get("machines", [])
        if len(machines) == 1:
            customer_machine_info[c["id"]] = machines[0].get("machine_name", "Unknown")
        else:
            customer_machine_info[c["id"]] = "Unknown"
    
    offline_machines = []
    
    for acc in customer_accounts:
        watcher_key = acc.get("watcher_key")
        if not watcher_key:
            continue
        customer = customer_map.get(acc.get("customer_id"), {})
        if customer.get("status") == "paused":
            continue
        
        machine_info = customer_machine_info.get(customer.get("id", ""), "Unknown")
        
        try:
            async with aiohttp.ClientSession() as session:
                url = f"https://www.viabtc.com/res/observer/worker?access_key={watcher_key}&coin=LTC"
                async with session.get(url, timeout=10) as resp:
                    data = await resp.json()
                    if data.get("code") == 0:
                        for w in data.get("data", {}).get("data", []):
                            status = w.get("worker_status", w.get("status", ""))
                            if status in ["offline", "unactive"]:
                                last_active = w.get("last_active", 0)
                                import time as t
                                offline_mins = int((t.time() - last_active) / 60) if last_active else 0
                                offline_machines.append({
                                    "name": w.get("worker_name", w.get("name", "")),
                                    "account": customer.get("name", acc.get("worker_name", "")),
                                    "machine_type": machine_info,
                                    "minutes_offline": offline_mins
                                })
        except:
            pass
    
    if not offline_machines:
        return {"success": True, "message": "All machines online! No alert needed.", "offline": 0}
    
    # Build alert details
    details = ""
    for m in offline_machines:
        hrs = m["minutes_offline"] // 60
        mins = m["minutes_offline"] % 60
        time_str = f"{hrs}h {mins}m" if hrs > 0 else f"{mins}m"
        details += f"❌ {m['name']} ({m['account']}) - offline {time_str}\n"
    
    try:
        phone = admin_phone if admin_phone.startswith("+") else "+" + admin_phone
        template_sid = os.environ.get("TWILIO_TEMPLATE_OFFLINE", "")
        if template_sid:
            import json
            msg = client.messages.create(
                content_sid=template_sid,
                content_variables=json.dumps({"1": str(len(offline_machines)), "2": details.strip()}),
                from_=get_whatsapp_from(),
                to=f"whatsapp:{phone}"
            )
        else:
            msg = client.messages.create(
                body=f"⚠️ WKBeast Alert: {len(offline_machines)} machine(s) OFFLINE\n\n{details}",
                from_=get_whatsapp_from(),
                to=f"whatsapp:{phone}"
            )
        return {"success": True, "sid": msg.sid, "offline": len(offline_machines), "machines": offline_machines}
    except Exception as e:
        return {"success": False, "error": str(e), "offline": len(offline_machines), "machines": offline_machines}

# ==================== BOT SETTINGS & AUTO-SCHEDULER ====================

@api_router.get("/bot-settings")
async def get_bot_settings():
    """Get WhatsApp bot automation settings"""
    settings = await db.bot_settings.find_one({"id": "bot_settings"}, {"_id": 0})
    if not settings:
        settings = {
            "id": "bot_settings",
            "auto_reminders_enabled": False,
            "offline_alerts_enabled": False,
            "admin_phone": "+9613022005",
            "reminder_day": 1,
            "reminder_interval_days": 3,
            "last_reminder_sent": None,
            "last_offline_check": None,
            "known_offline_workers": []
        }
        await db.bot_settings.insert_one(settings)
    settings.pop("_id", None)
    return settings

@api_router.put("/bot-settings")
async def update_bot_settings(data: dict):
    """Update WhatsApp bot automation settings"""
    allowed_keys = ["auto_reminders_enabled", "offline_alerts_enabled", "admin_phone", "reminder_day", "reminder_interval_days"]
    update_data = {k: v for k, v in data.items() if k in allowed_keys}
    
    result = await db.bot_settings.find_one_and_update(
        {"id": "bot_settings"},
        {"$set": update_data},
        upsert=True,
        return_document=True,
        projection={"_id": 0}
    )
    return result

@api_router.post("/bot/run-reminder-check")

@api_router.put("/customers/{customer_id}/whatsapp-toggle")
async def toggle_customer_whatsapp(customer_id: str):
    """Toggle WhatsApp notifications for a specific customer"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    current = customer.get("whatsapp_enabled", True)
    new_val = not current
    await db.customers.update_one({"id": customer_id}, {"$set": {"whatsapp_enabled": new_val}})
    return {"success": True, "whatsapp_enabled": new_val, "name": customer.get("name")}


async def run_reminder_check():
    """Manually trigger a payment reminder check (also runs automatically)"""
    bot_settings = await db.bot_settings.find_one({"id": "bot_settings"}, {"_id": 0})
    if not bot_settings or not bot_settings.get("auto_reminders_enabled"):
        return {"success": False, "message": "Auto reminders are disabled"}
    
    return await _send_auto_reminders()

@api_router.post("/bot/run-offline-check")
async def run_offline_check():
    """Manually trigger an offline machine check — ALWAYS sends alert"""
    bot_settings = await db.bot_settings.find_one({"id": "bot_settings"}, {"_id": 0})
    if not bot_settings or not bot_settings.get("offline_alerts_enabled"):
        return {"success": False, "message": "Offline alerts are disabled"}
    
    return await _check_offline_and_alert(force_send=True)

async def _send_auto_reminders():
    """Internal: Auto-generate payments if needed, then send WhatsApp reminders"""
    from datetime import datetime, timezone
    
    client_tw = get_twilio_client()
    if not client_tw:
        return {"success": False, "error": "Twilio not configured"}
    
    now = datetime.now(timezone.utc)
    current_month = now.strftime("%Y-%m")
    
    # Auto-generate payments for this month if not done yet
    active_customers = await db.customers.find({"status": "active"}, {"_id": 0}).to_list(1000)
    generated = 0
    for customer in active_customers:
        existing = await db.payments.find_one({"customer_id": customer["id"], "month": current_month})
        if not existing:
            prepaid = customer.get("prepaid_months", 0)
            status = "paid" if prepaid > 0 else "unpaid"
            payment = Payment(
                customer_id=customer["id"],
                customer_name=customer["name"],
                month=current_month,
                amount=customer["total_fee"],
                status=status
            )
            await db.payments.insert_one(payment.model_dump())
            if prepaid > 0:
                await db.customers.update_one({"id": customer["id"]}, {"$inc": {"prepaid_months": -1}})
            generated += 1
    
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    if not settings:
        return {"success": False, "error": "App settings not configured"}
    
    team = settings.get("team_name", "WKBeast Team")
    
    # On day 1: Send a nice message to prepaid customers
    prepaid_sent = 0
    if now.day == 1:
        all_customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
        for customer in all_customers:
            if customer.get("status") == "paused":
                continue
            if not customer.get("whatsapp_enabled", True):
                continue
            remaining = customer.get("prepaid_months", 0)
            # Customer just used a prepaid month (we decremented above), so add 1 back to show original
            # Only send if they HAD prepaid months this cycle (payment was auto-paid)
            payment = await db.payments.find_one({"customer_id": customer["id"], "month": current_month, "status": "paid"})
            if not payment:
                continue
            # Check if this was a prepaid payment (generated as paid this cycle or had prepaid)
            if remaining >= 0 and generated > 0:
                # Check if this specific customer's payment was just generated as paid
                was_prepaid = await db.payments.find_one({"customer_id": customer["id"], "month": current_month, "status": "paid"})
                if was_prepaid and remaining >= 0:
                    phone = customer.get("phone", "")
                    if not phone:
                        continue
                    phone_clean = format_phone_whatsapp(phone)
                    
                    if remaining == 0:
                        prepaid_msg = f"""✅ Dear {customer.get('name', 'Customer')},

Good news! Your {current_month} hosting fee has been automatically covered by your prepaid balance.

⚠️ This was your last prepaid month. Starting next month, you will need to pay your hosting fee on time.

Thank you for being ahead!
{team} 🐺💼

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""
                    elif remaining == 1:
                        prepaid_msg = f"""✅ Dear {customer.get('name', 'Customer')},

Good news! Your {current_month} hosting fee has been automatically covered by your prepaid balance.

📌 You have 1 month remaining on your prepaid balance. After that, monthly payments will resume.

Thank you for being ahead!
{team} 🐺💼

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""
                    else:
                        prepaid_msg = f"""✅ Dear {customer.get('name', 'Customer')},

Good news! Your {current_month} hosting fee has been automatically covered by your prepaid balance.

📌 You have {remaining} months remaining on your prepaid balance. Your next payment starts from {(now.month + remaining) % 12 or 12}/{now.year if now.month + remaining <= 12 else now.year + 1}.

Thank you for being ahead!
{team} 🐺💼

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""
                    
                    try:
                        client_tw.messages.create(body=prepaid_msg, from_=get_whatsapp_from(), to=f"whatsapp:{phone_clean}")
                        await db.whatsapp_logs.insert_one({
                            "customer_id": customer.get("id", ""),
                            "customer_name": customer.get("name", ""),
                            "phone": phone_clean,
                            "message": prepaid_msg,
                            "type": "prepaid_confirmation",
                            "direction": "outbound",
                            "status": "sent",
                            "created_at": now.isoformat()
                        })
                        prepaid_sent += 1
                    except:
                        pass
    
    # Now send reminders to unpaid customers
    payments = await db.payments.find({"month": current_month, "status": "unpaid"}, {"_id": 0}).to_list(10000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    customer_map = {c["id"]: c for c in customers}
    
    day = now.day
    
    # Determine message based on day
    whish = settings.get("whish_number", "03022005")
    usdt = settings.get("usdt_address", "")
    team = settings.get("team_name", "WKBeast Team")
    
    sent = 0
    failed = 0
    message_type = "day1"
    
    for payment in payments:
        customer = customer_map.get(payment.get("customer_id"))
        if not customer or customer.get("status") == "paused":
            continue
        
        if not customer.get("whatsapp_enabled", True):
            continue
        
        phone = customer.get("phone", "")
        if not phone:
            continue
        
        phone_clean = format_phone_whatsapp(phone)
        amount = payment.get("amount", 0)
        
        # Day 1: Friendly initial reminder
        if day == 1:
            message_type = "initial"
            message = f"""📢 Dear Valued Customer,

This is a kind reminder to please settle your hosting fees before the 2nd of each month, as we have major financial obligations to cover by that date.

Unlike other farms, we don't request payments 6 months in advance — but we kindly ask for your cooperation in making the payment on time each month.

🔹 {current_month} Hosting Fee: ${amount}
🔹 Payment Options:
• Whish: {whish}
• USDT (BEP20 Network): {usdt}

Thank you for your continued trust and support. Your timely payment helps us keep everything running smoothly.

Warm regards,
{team} 🐺💼

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""

        # Day 2: Reminder
        elif day == 2:
            message_type = "reminder"
            message = f"""⏰ Friendly Reminder

Hi! Just a quick follow-up — your hosting fee of ${amount} for {current_month} is still pending.

Please settle it at your earliest convenience:
• Whish: {whish}
• USDT (BEP20): {usdt}

Thank you!
{team}

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""

        # Day 3: Final warning
        elif day == 3:
            message_type = "warning"
            message = f"""⚠️ Important Notice

Dear customer, your hosting fee of ${amount} for {current_month} is overdue.

Please be advised that machines with unpaid fees will be powered off tonight at midnight.

To avoid service interruption, please pay now:
• Whish: {whish}
• USDT (BEP20): {usdt}

If you have already paid, please ignore this message.

{team}

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""

        # Day 4+: Short reminder every 3 days
        else:
            message_type = "follow_up"
            message = f"""📌 Payment still pending: ${amount} for {current_month}

• Whish: {whish}
• USDT (BEP20): {usdt}

{team}

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""
        
        try:
            client_tw.messages.create(
                body=message,
                from_=get_whatsapp_from(),
                to=f"whatsapp:{phone_clean}"
            )
            # Log message to DB
            await db.whatsapp_logs.insert_one({
                "customer_id": payment.get("customer_id", ""),
                "customer_name": customer.get("name", ""),
                "phone": phone_clean,
                "message": message,
                "type": message_type,
                "direction": "outbound",
                "status": "sent",
                "created_at": now.isoformat()
            })
            sent += 1
        except Exception as e:
            await db.whatsapp_logs.insert_one({
                "customer_id": payment.get("customer_id", ""),
                "customer_name": customer.get("name", ""),
                "phone": phone_clean,
                "message": message,
                "type": message_type,
                "direction": "outbound",
                "status": f"failed: {str(e)}",
                "created_at": now.isoformat()
            })
            failed += 1
    
    # Update last sent timestamp
    await db.bot_settings.update_one(
        {"id": "bot_settings"},
        {"$set": {"last_reminder_sent": now.isoformat()}}
    )
    
    return {"success": True, "sent": sent, "failed": failed, "generated": generated, "prepaid_notified": prepaid_sent, "month": current_month, "message_type": message_type}

async def _check_offline_and_alert(force_send=False):
    """Internal: Check for offline machines and alert admin. force_send=True always sends alert."""
    import aiohttp
    from datetime import datetime, timezone
    
    client_tw = get_twilio_client()
    if not client_tw:
        return {"success": False, "error": "Twilio not configured"}
    
    bot_settings = await db.bot_settings.find_one({"id": "bot_settings"}, {"_id": 0})
    admin_phone = bot_settings.get("admin_phone", "+9613022005") if bot_settings else "+9613022005"
    known_offline = set(bot_settings.get("known_offline_workers", [])) if bot_settings else set()
    
    # Fetch current machine status using watcher keys
    customer_accounts = await db.customer_accounts.find({}, {"_id": 0}).to_list(1000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    customer_map = {c["id"]: c for c in customers}
    settings = await db.viabtc_settings.find_one({"id": "viabtc_settings"}, {"_id": 0})
    main_watcher_key = settings.get("watcher_key") if settings else None
    
    # Collect all unique watcher keys to check
    watcher_accounts = []
    seen_keys = set()
    
    for acc in customer_accounts:
        wk = acc.get("watcher_key")
        if wk and wk not in seen_keys:
            seen_keys.add(wk)
            customer = customer_map.get(acc.get("customer_id"), {})
            if customer.get("status") != "paused":
                watcher_accounts.append({
                    "watcher_key": wk,
                    "worker_name": acc.get("worker_name", ""),
                    "display_name": customer.get("name", acc.get("worker_name", ""))
                })
    
    # Add main watcher key if not already included
    if main_watcher_key and main_watcher_key not in seen_keys:
        watcher_accounts.append({
            "watcher_key": main_watcher_key,
            "worker_name": "turkbeast",
            "display_name": "turkbeast (Main)"
        })
    
    # Build machine type lookup per customer from DB
    customer_machine_info = {}
    for c in customers:
        machines = c.get("machines", [])
        if len(machines) == 1:
            customer_machine_info[c["id"]] = machines[0].get("machine_name", "Unknown")
        else:
            customer_machine_info[c["id"]] = "Unknown"
    
    # Also map by worker_name for customer_accounts lookup
    account_to_customer_id = {}
    for acc in customer_accounts:
        account_to_customer_id[acc.get("worker_name", "")] = acc.get("customer_id", "")
    
    # Get machine IPs from live data (pushed by PC sync)
    live_machines = await db.machine_live_data.find({}, {"_id": 0, "worker_name": 1, "ip": 1}).to_list(10000)
    worker_to_ip = {}
    for lm in live_machines:
        wn = lm.get("worker_name", "").lower().strip()
        if wn and lm.get("ip"):
            worker_to_ip[wn] = lm["ip"]
    
    current_offline = set()
    offline_details = []
    
    async with aiohttp.ClientSession() as session:
        for acc in watcher_accounts:
            watcher_key = acc["watcher_key"]
            display_name = acc["display_name"]
            worker_name = acc.get("worker_name", "")
            
            # Get customer's machine info
            cust_id = account_to_customer_id.get(worker_name, "")
            machine_info = customer_machine_info.get(cust_id, "Unknown")
            
            try:
                url = f"https://www.viabtc.com/res/observer/worker?access_key={watcher_key}&coin=LTC"
                async with session.get(url, timeout=10) as resp:
                    data = await resp.json()
                    if data.get("code") == 0:
                        for w in data.get("data", {}).get("data", []):
                            wname = w.get("worker_name", w.get("name", ""))
                            status = w.get("worker_status", w.get("status", ""))
                            worker_id = f"{worker_name}:{wname}"
                            
                            if status in ["offline", "unactive"]:
                                current_offline.add(worker_id)
                                last_active = w.get("last_active", 0)
                                
                                # Look up IP from live data
                                machine_ip = worker_to_ip.get(wname.lower(), worker_to_ip.get(worker_name.lower(), ""))
                                
                                import time as t
                                offline_mins = int((t.time() - last_active) / 60) if last_active else 0
                                offline_details.append({
                                    "worker_id": worker_id,
                                    "name": wname,
                                    "account": display_name,
                                    "machine_type": machine_info,
                                    "machine_ip": machine_ip,
                                    "minutes_offline": offline_mins
                                })
            except:
                pass
    
    # Determine newly offline and newly back online
    if force_send:
        # Manual check: always report ALL offline machines
        newly_offline = current_offline
        back_online = set()
    else:
        # Auto check: only report changes
        newly_offline = current_offline - known_offline
        back_online = known_offline - current_offline
    
    messages_sent = []
    template_sid = os.environ.get("TWILIO_TEMPLATE_OFFLINE", "")
    
    # Alert for newly offline machines
    if newly_offline:
        details = ""
        for worker_id in newly_offline:
            detail = next((d for d in offline_details if d["worker_id"] == worker_id), None)
            if detail:
                ip_str = f" | IP: {detail['machine_ip']}" if detail.get('machine_ip') else ""
                details += f"❌ {detail['name']} ({detail['account']}){ip_str}\n"
            else:
                details += f"❌ {worker_id}\n"
        
        try:
            import json
            phone = admin_phone if admin_phone.startswith("+") else "+" + admin_phone
            if template_sid:
                client_tw.messages.create(
                    content_sid=template_sid,
                    content_variables=json.dumps({"1": str(len(newly_offline)), "2": details.strip()}),
                    from_=get_whatsapp_from(), to=f"whatsapp:{phone}"
                )
            else:
                client_tw.messages.create(body=f"⚠️ ALERT: {len(newly_offline)} machine(s) went OFFLINE\n\n{details}", from_=get_whatsapp_from(), to=f"whatsapp:{phone}")
            messages_sent.append("offline_alert")
        except:
            pass
    
    # Alert for machines back online
    if back_online:
        recovery_details = ""
        for worker_id in back_online:
            recovery_details += f"✅ {worker_id}\n"
        
        try:
            import json
            phone = admin_phone if admin_phone.startswith("+") else "+" + admin_phone
            if template_sid:
                client_tw.messages.create(
                    content_sid=template_sid,
                    content_variables=json.dumps({"1": "0", "2": f"✅ RECOVERED: {len(back_online)} machine(s) back ONLINE\n{recovery_details.strip()}"}),
                    from_=get_whatsapp_from(), to=f"whatsapp:{phone}"
                )
            else:
                client_tw.messages.create(body=f"✅ RECOVERED: {len(back_online)} machine(s) back ONLINE\n\n{recovery_details}", from_=get_whatsapp_from(), to=f"whatsapp:{phone}")
            messages_sent.append("recovery_alert")
        except:
            pass
    
    # Update known offline list
    await db.bot_settings.update_one(
        {"id": "bot_settings"},
        {"$set": {
            "known_offline_workers": list(current_offline),
            "last_offline_check": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "success": True,
        "current_offline": len(current_offline),
        "newly_offline": len(newly_offline),
        "back_online": len(back_online),
        "messages_sent": messages_sent
    }


# ==================== WHATSAPP WEBHOOK (Receive incoming messages) ====================

@api_router.post("/whatsapp/webhook")
async def whatsapp_webhook(request: Request):
    """Receive incoming WhatsApp messages from Twilio webhook"""
    form_data = await request.form()
    
    from_number = form_data.get("From", "").replace("whatsapp:", "")
    body = form_data.get("Body", "")
    message_sid = form_data.get("MessageSid", "")
    
    # Find customer by phone number
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    matched_customer = None
    for c in customers:
        customer_phone = format_phone_whatsapp(c.get("phone", ""))
        if customer_phone == from_number:
            matched_customer = c
            break
    
    from datetime import datetime, timezone
    
    # Log the incoming message
    await db.whatsapp_logs.insert_one({
        "customer_id": matched_customer.get("id", "") if matched_customer else "",
        "customer_name": matched_customer.get("name", "Unknown") if matched_customer else f"Unknown ({from_number})",
        "phone": from_number,
        "message": body,
        "type": "incoming",
        "direction": "inbound",
        "status": "received",
        "message_sid": message_sid,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Return empty TwiML (no auto-reply)
    return Response(content="<Response></Response>", media_type="application/xml")

# ==================== PAYMENT CONFIRMATION NOTIFICATION ====================


@api_router.post("/whatsapp/broadcast")
async def broadcast_whatsapp(data: dict):
    """Send a message to ALL active customers"""
    client_tw = get_twilio_client()
    if not client_tw:
        return {"success": False, "error": "Twilio not configured"}
    
    message = data.get("message", "")
    if not message.strip():
        return {"success": False, "error": "Message cannot be empty"}
    
    # Add bot disclaimer
    full_message = f"""{message}

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""
    
    customers = await db.customers.find({"status": {"$ne": "paused"}}, {"_id": 0}).to_list(10000)
    
    from datetime import datetime, timezone
    sent = 0
    failed = 0
    
    for customer in customers:
        phone = customer.get("phone", "")
        if not phone:
            continue
        
        if not customer.get("whatsapp_enabled", True):
            continue
        
        phone_clean = format_phone_whatsapp(phone)
        
        try:
            client_tw.messages.create(
                body=full_message,
                from_=get_whatsapp_from(),
                to=f"whatsapp:{phone_clean}"
            )
            await db.whatsapp_logs.insert_one({
                "customer_id": customer.get("id", ""),
                "customer_name": customer.get("name", ""),
                "phone": phone_clean,
                "message": full_message,
                "type": "broadcast",
                "direction": "outbound",
                "status": "sent",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            sent += 1
        except Exception as e:
            await db.whatsapp_logs.insert_one({
                "customer_id": customer.get("id", ""),
                "customer_name": customer.get("name", ""),
                "phone": phone_clean,
                "message": full_message,
                "type": "broadcast",
                "direction": "outbound",
                "status": f"failed: {str(e)}",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
            failed += 1
    
    return {"success": True, "sent": sent, "failed": failed, "total_customers": len(customers)}


@api_router.post("/whatsapp/send-payment-confirmation")
async def send_payment_confirmation(customer_id: str, month: str):
    """Send payment confirmation message to customer"""
    client_tw = get_twilio_client()
    if not client_tw:
        return {"success": False, "error": "Twilio not configured"}
    
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        return {"success": False, "error": "Customer not found"}
    
    phone = customer.get("phone", "")
    if not phone:
        return {"success": False, "error": "Customer has no phone number"}
    
    phone_clean = format_phone_whatsapp(phone)
    settings = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    team = settings.get("team_name", "WKBeast Team") if settings else "WKBeast Team"
    
    message = f"""✅ Payment Confirmed

Dear {customer.get('name', 'Customer')},

Thank you! We have reviewed and confirmed your payment for {month}.

Your machines will continue to run without interruption. We appreciate your timely payment.

Best regards,
{team} 🐺💼

━━━━━━━━━━━━━━━
🤖 This is an automated message. Please do not reply to this number. For questions, contact us directly."""
    
    from datetime import datetime, timezone
    
    try:
        msg = client_tw.messages.create(
            body=message,
            from_=get_whatsapp_from(),
            to=f"whatsapp:{phone_clean}"
        )
        # Log it
        await db.whatsapp_logs.insert_one({
            "customer_id": customer_id,
            "customer_name": customer.get("name", ""),
            "phone": phone_clean,
            "message": message,
            "type": "payment_confirmed",
            "direction": "outbound",
            "status": "sent",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        return {"success": True, "sid": msg.sid}
    except Exception as e:
        return {"success": False, "error": str(e)}



# ==================== WHATSAPP MESSAGE HISTORY ====================

@api_router.get("/whatsapp/history")
async def get_whatsapp_history(customer_id: str = None, limit: int = 100):
    """Get WhatsApp message history, optionally filtered by customer"""
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    
    logs = await db.whatsapp_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return logs

@api_router.get("/whatsapp/history/all")
async def get_all_whatsapp_history():
    """Get all WhatsApp message history grouped by customer"""
    logs = await db.whatsapp_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Group by customer
    by_customer = {}
    for log in logs:
        cname = log.get("customer_name", "Unknown")
        if cname not in by_customer:
            by_customer[cname] = {"customer_name": cname, "customer_id": log.get("customer_id", ""), "phone": log.get("phone", ""), "messages": []}
        by_customer[cname]["messages"].append(log)
    
    return list(by_customer.values())


# ==================== BACKGROUND SCHEDULER ====================

_scheduler_running = False

async def bot_scheduler():
    """Background task that runs payment reminders and offline checks"""
    global _scheduler_running
    _scheduler_running = True
    
    while _scheduler_running:
        try:
            from datetime import datetime, timezone
            now = datetime.now(timezone.utc)
            
            bot_settings = await db.bot_settings.find_one({"id": "bot_settings"}, {"_id": 0})
            if not bot_settings:
                await asyncio.sleep(60)
                continue
            
            # --- PAYMENT REMINDERS ---
            if bot_settings.get("auto_reminders_enabled"):
                last_sent = bot_settings.get("last_reminder_sent")
                
                should_send = False
                # Days 1, 2, 3: send every day (escalating messages)
                if now.day in [1, 2, 3]:
                    if not last_sent or last_sent[:10] != now.strftime("%Y-%m-%d"):
                        should_send = True
                # Day 4+: send every 3 days
                elif now.day > 3:
                    if last_sent:
                        last_dt = datetime.fromisoformat(last_sent.replace("Z", "+00:00")) if isinstance(last_sent, str) else last_sent
                        days_since = (now - last_dt).days
                        if days_since >= 3:
                            should_send = True
                    else:
                        should_send = True
                
                if should_send:
                    await _send_auto_reminders()
            
            # --- OFFLINE ALERTS (check every 10 minutes) ---
            if bot_settings.get("offline_alerts_enabled"):
                last_check = bot_settings.get("last_offline_check")
                should_check = True
                if last_check:
                    last_dt = datetime.fromisoformat(last_check.replace("Z", "+00:00")) if isinstance(last_check, str) else last_check
                    mins_since = (now - last_dt).total_seconds() / 60
                    if mins_since < 10:
                        should_check = False
                
                if should_check:
                    await _check_offline_and_alert()
            
        except Exception as e:
            logging.error(f"Bot scheduler error: {e}")
        
        # Sleep 60 seconds between checks
        await asyncio.sleep(60)

@app.on_event("startup")
async def start_scheduler():
    asyncio.create_task(bot_scheduler())

# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
